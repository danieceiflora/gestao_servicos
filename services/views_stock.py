from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.db.models import Q, Sum, F, ExpressionWrapper, DecimalField, Count
from django.db.models.functions import Coalesce
from django.forms import inlineformset_factory
import csv
import io
import hashlib
from decimal import Decimal
from openpyxl import load_workbook
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from .models import (
    Product, StockMovement, ImportHistory, ImportItem, ProductCategory,
    Supplier, PurchaseInvoice, PurchaseInvoiceItem,
)
from integracoes.models import SystemConfig
from .forms import (
    ProductForm, StockMovementForm, ProductImportForm, ProductCompositionFormSet, ProductVariantFormSet,
    PurchaseInvoiceForm, PurchaseInvoiceItemForm, PurchaseInvoiceItemFormSet, PurchaseInvoiceXMLUploadForm,
)
from .utils_nfe_import import (
    parse_nfe_xml, match_products, NFeXMLParseError,
    find_supplier_by_cnpj, get_or_create_supplier_from_xml, create_missing_products,
    _update_product_fiscal,
)

class ManagerRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.is_manager

def is_manager(user):
    return user.is_authenticated and user.is_manager

@login_required
def get_product_prices(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return JsonResponse({
        'preco_custo': float(product.preco_custo),
        'default_unit_price': float(product.default_unit_price),
        'current_stock': float(product.current_stock),
        'min_stock': float(product.min_stock),
        'max_stock': float(product.max_stock),
        'unit_type': product.get_unit_type_display(),
    })

@login_required
def search_materia_prima(request):
    q = request.GET.get('q', '')
    products = Product.objects.filter(
        Q(name__icontains=q) | Q(code__icontains=q),
        type=Product.Type.MATERIA_PRIMA,
        is_active=True
    )[:10]
    
    results = [
        {'id': p.id, 'text': f"{p.name} ({p.code or 'S/C'})"} 
        for p in products
    ]
    return JsonResponse({'results': results})

class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'services/product_list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        queryset = Product.objects.select_related('category').all()
        search = self.request.GET.get('search')
        category = self.request.GET.get('category')
        status = self.request.GET.get('status')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search)
            )
        if category:
            queryset = queryset.filter(category_id=category)
        if status == 'low':
            queryset = queryset.filter(current_stock__gt=0, current_stock__lte=F('min_stock'))
        elif status == 'zero':
            queryset = queryset.filter(current_stock__lte=0)
        elif status == 'ok':
            queryset = queryset.filter(current_stock__gt=F('min_stock'))
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = Product.objects.all()
        valor_total = qs.aggregate(
            total=Coalesce(Sum(ExpressionWrapper(
                F('current_stock') * F('preco_custo'),
                output_field=DecimalField()
            )), Decimal('0'))
        )['total']
        context['kpi_valor_total'] = valor_total
        context['kpi_total_skus'] = qs.count()
        context['kpi_abaixo_minimo'] = qs.filter(current_stock__gt=0, current_stock__lte=F('min_stock'), min_stock__gt=0).count()
        context['kpi_zerados'] = qs.filter(current_stock__lte=0).count()
        context['categories'] = ProductCategory.objects.all().order_by('name')
        context['selected_category'] = self.request.GET.get('category', '')
        context['selected_status'] = self.request.GET.get('status', '')
        return context

class ProductCreateView(LoginRequiredMixin, ManagerRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'services/product_form.html'
    success_url = reverse_lazy('product_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tax_regime'] = SystemConfig.load().tax_regime
        if self.request.POST:
            context['composition_formset'] = ProductCompositionFormSet(self.request.POST)
            context['variant_formset'] = ProductVariantFormSet(self.request.POST)
        else:
            context['composition_formset'] = ProductCompositionFormSet()
            context['variant_formset'] = ProductVariantFormSet()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        composition_formset = context['composition_formset']
        variant_formset = context['variant_formset']

        with transaction.atomic():
            self.object = form.save()
            if form.instance.format == 'COM_COMPOSICAO':
                if composition_formset.is_valid():
                    composition_formset.instance = self.object
                    composition_formset.save()
                else:
                    return self.form_invalid(form)
            elif form.instance.format == 'COM_VARIACOES':
                if variant_formset.is_valid():
                    variant_formset.instance = self.object
                    variant_formset.save()
                else:
                    return self.form_invalid(form)

        messages.success(self.request, "Produto cadastrado com sucesso.")
        return redirect(self.success_url)

class ProductUpdateView(LoginRequiredMixin, ManagerRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'services/product_form.html'
    success_url = reverse_lazy('product_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tax_regime'] = SystemConfig.load().tax_regime
        if self.request.POST:
            context['composition_formset'] = ProductCompositionFormSet(self.request.POST, instance=self.object)
            context['variant_formset'] = ProductVariantFormSet(self.request.POST, instance=self.object)
        else:
            context['composition_formset'] = ProductCompositionFormSet(instance=self.object)
            context['variant_formset'] = ProductVariantFormSet(instance=self.object)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        composition_formset = context['composition_formset']
        variant_formset = context['variant_formset']

        with transaction.atomic():
            self.object = form.save()
            if form.instance.format == 'COM_COMPOSICAO':
                if composition_formset.is_valid():
                    composition_formset.instance = self.object
                    composition_formset.save()
                else:
                    return self.form_invalid(form)
            elif form.instance.format == 'COM_VARIACOES':
                if variant_formset.is_valid():
                    variant_formset.instance = self.object
                    variant_formset.save()
                else:
                    return self.form_invalid(form)

        messages.success(self.request, "Produto atualizado com sucesso.")
        return redirect(self.success_url)

class StockMovementCreateView(LoginRequiredMixin, ManagerRequiredMixin, CreateView):
    model = StockMovement
    form_class = StockMovementForm
    template_name = 'services/stock_movement_form.html'
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        movement = form.save(commit=False)
        product = movement.product
        
        if movement.movement_type == StockMovement.MovementType.ENTRADA:
            product.increase_stock(
                movement.quantity,
                user=self.request.user,
                reason=movement.reason,
                notes=movement.notes,
                service_order=movement.service_order
            )
        else:
            product.reduce_stock(
                movement.quantity,
                user=self.request.user,
                reason=movement.reason,
                notes=movement.notes,
                service_order=movement.service_order
            )
        
        messages.success(self.request, f"Movimentação de {movement.get_movement_type_display()} realizada com sucesso.")
        return redirect(self.success_url)

@login_required
def product_stock_history(request, pk):
    from datetime import datetime, date
    product = get_object_or_404(Product, pk=pk)
    movements = product.movements.all().select_related('user', 'service_order')

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    mov_type = request.GET.get('type')

    if date_from:
        try:
            movements = movements.filter(created_at__date__gte=date_from)
        except (ValueError, TypeError):
            pass
    if date_to:
        try:
            movements = movements.filter(created_at__date__lte=date_to)
        except (ValueError, TypeError):
            pass
    if mov_type in ('ENTRADA', 'SAIDA'):
        movements = movements.filter(movement_type=mov_type)

    # Dados para gráfico (últimos 30 lançamentos cronológicos)
    chart_qs = list(product.movements.order_by('created_at').values(
        'created_at', 'movement_type', 'quantity', 'saldo_apos'
    )[:60])
    chart_labels = [m['created_at'].strftime('%d/%m %H:%M') for m in chart_qs]
    chart_data = [float(m['saldo_apos']) if m['saldo_apos'] is not None else None for m in chart_qs]

    context = {
        'product': product,
        'movements': movements,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'selected_type': mov_type or '',
        'chart_labels': chart_labels,
        'chart_data': chart_data,
    }
    return render(request, 'services/product_stock_history.html', context)

@login_required
@user_passes_test(is_manager)
def product_import_template(request):
    """Gera um arquivo modelo (CSV ou XLSX) para importação, baseada no tipo de operação."""
    file_format = request.GET.get('format', 'csv')
    op_type = request.GET.get('type', 'CATALOG')
    
    if op_type == 'STOCK':
        headers = ['codigo', 'quantidade', 'tipo_entrada_saida', 'motivo', 'notas']
        data = [
            ['CABO-25-PT', '50', 'ENTRADA', 'COMPRA', 'NF 1234'],
            ['DISJ-20A', '2', 'SAIDA', 'PERDA', 'Quebrado no transporte']
        ]
        filename_prefix = "modelo_movimentacao_estoque"
    else:
        headers = ['nome', 'codigo', 'unidade', 'preco_venda', 'ativo_sim_nao']
        data = [
            ['Cabo Flexível 2.5mm', 'CABO-25-PT', 'M', '4.50', 'S'],
            ['Disjuntor 20A', 'DISJ-20A', 'UN', '18.90', 'S']
        ]
        filename_prefix = "modelo_gestao_catalogo"

    if file_format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo Importação"
        
        # Header styling
        header_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add sample data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.xlsx"'
        wb.save(response)
        return response
    else:
        # Default to CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in data:
            writer.writerow(row)
        
        return response

@login_required
@user_passes_test(is_manager)
def product_import(request):
    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            op_type = form.cleaned_data['operation_type']
            file = request.FILES['file']
            filename = file.name
            
            # --- PREVENÇÃO DE DUPLICIDADE (CHECKSUM) ---
            file_content = file.read()
            file_hash = hashlib.sha256(file_content).hexdigest()
            
            if ImportHistory.objects.filter(file_hash=file_hash).exists():
                messages.error(request, "Este arquivo (ou um conteúdo idêntico) já foi importado anteriormente. Operação cancelada para evitar duplicidade.")
                return redirect('product_import')
            
            # Reset file pointer for processing after reading for hash
            file.seek(0)
            
            data = []
            try:
                if filename.endswith('.csv'):
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    data = list(reader)
                elif filename.endswith('.xlsx'):
                    wb = load_workbook(file)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        data.append(dict(zip(headers, row)))
            except Exception as e:
                messages.error(request, f"Erro ao ler o arquivo: {str(e)}")
                return redirect('product_import')

            unit_map = {
                'UN': Product.UnitType.UNIT,
                'M': Product.UnitType.METER,
                'M2': Product.UnitType.SQUARE_METER,
                'L': Product.UnitType.LITER,
                'KG': Product.UnitType.KILO,
            }

            # Mapping for movement reasons
            reason_map = {
                'COMPRA': StockMovement.Reason.PURCHASE,
                'VENDA': StockMovement.Reason.SALE_OS,
                'AJUSTE': StockMovement.Reason.ADJUSTMENT,
                'PERDA': StockMovement.Reason.LOSS,
                'DEVOLUCAO': StockMovement.Reason.RETURN,
            }

            created_count = 0
            updated_count = 0
            movements_count = 0
            errors = []

            with transaction.atomic():
                audit = ImportHistory.objects.create(
                    user=request.user,
                    filename=f"[{op_type}] {filename}",
                    file_hash=file_hash
                )

                for index, row in enumerate(data, start=2):
                    try:
                        code = row.get('codigo')
                        name = row.get('nome')
                        identifier = code or name or f"Linha {index}"
                        
                        if not code and op_type == 'STOCK':
                            ImportItem.objects.create(
                                import_history=audit, row_number=index, identifier=identifier,
                                action="Falha", details="Código/SKU é obrigatório para estoque.", is_error=True
                            )
                            errors.append(f"Linha {index}: Código obrigatório.")
                            continue

                        if op_type == 'STOCK':
                            # --- OPERAÇÃO: MOVIMENTAÇÃO DE ESTOQUE ---
                            try:
                                product = Product.objects.get(code=code)
                            except Product.DoesNotExist:
                                ImportItem.objects.create(
                                    import_history=audit, row_number=index, identifier=identifier,
                                    action="Falha", details=f"Produto com código {code} não encontrado.", is_error=True
                                )
                                errors.append(f"Linha {index}: Produto não encontrado.")
                                continue

                            try:
                                qty = float(str(row.get('quantidade') or 0).replace(',', '.'))
                                if qty <= 0: raise ValueError()
                            except (ValueError, TypeError):
                                ImportItem.objects.create(
                                    import_history=audit, row_number=index, identifier=identifier,
                                    product=product, action="Falha", details="Quantidade inválida.", is_error=True
                                )
                                errors.append(f"Linha {index}: Qtd inválida.")
                                continue

                            tipo_str = str(row.get('tipo_entrada_saida') or 'ENTRADA').upper().strip()
                            mov_type = StockMovement.MovementType.ENTRADA if tipo_str in ['ENTRADA', 'E'] else StockMovement.MovementType.SAIDA
                            
                            reason_str = str(row.get('motivo') or 'AJUSTE').upper().strip()
                            reason = reason_map.get(reason_str, StockMovement.Reason.ADJUSTMENT)
                            
                            notes = row.get('notas') or f"Importação em massa ({filename})"

                            old_stock = float(product.current_stock)
                            if mov_type == StockMovement.MovementType.ENTRADA:
                                product.increase_stock(
                                    Decimal(str(qty)),
                                    user=request.user,
                                    reason=reason,
                                    notes=notes
                                )
                            else:
                                product.reduce_stock(
                                    Decimal(str(qty)),
                                    user=request.user,
                                    reason=reason,
                                    notes=notes
                                )

                            ImportItem.objects.create(
                                import_history=audit, row_number=index, identifier=identifier,
                                product=product, action="Estoque", 
                                details=f"{'Acrescentado' if mov_type == StockMovement.MovementType.ENTRADA else 'Removido'} {qty} ({reason_str}). Saldo: {old_stock} -> {product.current_stock}"
                            )
                            movements_count += 1
                            updated_count += 1

                        else:
                            # --- OPERAÇÃO: GESTÃO DE CATÁLOGO ---
                            if not name:
                                ImportItem.objects.create(
                                    import_history=audit, row_number=index, identifier=identifier,
                                    action="Falha", details="Nome obrigatório para catálogo.", is_error=True
                                )
                                continue
                            
                            unit_str = str(row.get('unidade') or 'UN').upper().strip()
                            unit_type = unit_map.get(unit_str, Product.UnitType.UNIT)
                            
                            try:
                                price = float(str(row.get('preco_venda') or 0).replace(',', '.'))
                            except (ValueError, TypeError): price = 0
                            
                            ativo_str = str(row.get('ativo_sim_nao') or 'S').upper().strip()
                            is_active = ativo_str in ['S', 'SIM', 'Y', 'YES', '1', 'TRUE']

                            product, created = Product.objects.get_or_create(
                                code=code,
                                defaults={'name': name, 'unit_type': unit_type, 'default_unit_price': price, 'is_active': is_active}
                            )

                            if created:
                                ImportItem.objects.create(
                                    import_history=audit, row_number=index, identifier=identifier,
                                    product=product, action="Criado", details=f"Produto novo. Preço: {price}. Ativo: {is_active}"
                                )
                                created_count += 1
                            else:
                                changes = []
                                if product.name != name: changes.append(f"Nome: {product.name} -> {name}")
                                if float(product.default_unit_price) != price: changes.append(f"Preço: {product.default_unit_price} -> {price}")
                                if product.is_active != is_active: changes.append(f"Ativo: {product.is_active} -> {is_active}")
                                
                                product.name = name
                                product.unit_type = unit_type
                                product.default_unit_price = price
                                product.is_active = is_active
                                product.save()
                                
                                ImportItem.objects.create(
                                    import_history=audit, row_number=index, identifier=identifier,
                                    product=product, action="Atualizado", details=", ".join(changes) if changes else "Nenhuma alteração detectada"
                                )
                                updated_count += 1

                    except Exception as e:
                        ImportItem.objects.create(
                            import_history=audit, row_number=index, identifier=f"Linha {index}",
                            action="Erro Crítico", details=str(e), is_error=True
                        )
                        errors.append(f"Linha {index}: {str(e)}")

                audit.created_count = created_count
                audit.updated_count = updated_count
                audit.movements_count = movements_count
                if errors:
                    audit.errors = "\n".join(errors)
                audit.save()

            if errors:
                messages.warning(request, f"Importação [{op_type}] finalizada com erros nas linhas: {', '.join(errors[:3])}...")
            
            messages.success(request, f"Importação [{op_type}] concluída. {created_count} criados, {updated_count} atualizados.")
            return redirect('product_list')
    else:
        form = ProductImportForm()
    
    return render(request, 'services/product_import.html', {'form': form})

class ImportHistoryListView(LoginRequiredMixin, ManagerRequiredMixin, ListView):
    model = ImportHistory
    template_name = 'services/product_import_history.html'
    context_object_name = 'imports'
    paginate_by = 20

class ImportHistoryDetailView(LoginRequiredMixin, ManagerRequiredMixin, DetailView):
    model = ImportHistory
    template_name = 'services/product_import_history_detail.html'
    context_object_name = 'import_obj'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items_logged.all().select_related('product')
        return context


# --- NOTA DE ENTRADA (PURCHASE INVOICE) ---

def _item_from_hidden_post(post_data, index):
    return {
        'code': post_data.get(f'item_{index}_code', ''),
        'barcode': post_data.get(f'item_{index}_barcode', ''),
        'name': post_data.get(f'item_{index}_name', ''),
        'unit': post_data.get(f'item_{index}_unit', ''),
        'ncm': post_data.get(f'item_{index}_ncm', ''),
        'cfop': post_data.get(f'item_{index}_cfop', ''),
        'cest': post_data.get(f'item_{index}_cest', ''),
        'quantity': _decimal_or_zero(post_data.get(f'item_{index}_quantity')),
        'unit_cost': _decimal_or_zero(post_data.get(f'item_{index}_unit_cost')),
        'origem': _decimal_or_zero(post_data.get(f'item_{index}_origem')),
        'icms_raw': post_data.get(f'item_{index}_icms_raw', ''),
        'pis_cst': post_data.get(f'item_{index}_pis_cst', ''),
        'cofins_cst': post_data.get(f'item_{index}_cofins_cst', ''),
    }


def _decimal_or_zero(value):
    if value in (None, ''):
        return Decimal('0')
    try:
        # Aceita tanto "25.00" quanto "25,00" — proteção extra caso algum
        # número chegue localizado (ex: template renderizando em pt-BR).
        return Decimal(str(value).replace(',', '.'))
    except Exception:
        return Decimal('0')


@login_required
@user_passes_test(is_manager)
def purchase_invoice_import_preview(request):
    """Tela única de importação: passo 1 lê o XML e mostra os dados extraídos
    (nota, fornecedor, itens) pra conferência; passo 2 ("Confirmar e
    Importar") efetivamente cria a Nota de Entrada (RASCUNHO) com esses
    dados, cadastrando fornecedor/produtos que não existirem. O usuário
    revisa e lança (dá entrada no estoque) na tela de detalhe da nota."""
    parsed = None
    matched_items = None
    supplier = None
    total = None
    xml_import_error = ''
    raw_preview = ''
    filename = ''
    duplicate_invoice = None

    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        try:
            item_count = int(request.POST.get('item_count', 0))
        except (TypeError, ValueError):
            item_count = 0

        parsed = {
            'number': request.POST.get('h_number', ''),
            'series': request.POST.get('h_series', ''),
            'access_key': request.POST.get('h_access_key', ''),
            'issue_date': request.POST.get('h_issue_date', ''),
            'supplier_cnpj': request.POST.get('h_supplier_cnpj', ''),
            'supplier_name': request.POST.get('h_supplier_name', ''),
            'supplier_trade_name': request.POST.get('h_supplier_trade_name', ''),
            'supplier_ie': request.POST.get('h_supplier_ie', ''),
            'supplier_phone': request.POST.get('h_supplier_phone', ''),
            'supplier_address': request.POST.get('h_supplier_address', ''),
            'items': [_item_from_hidden_post(request.POST, i) for i in range(item_count)],
        }

        supplier, supplier_created = get_or_create_supplier_from_xml(parsed)
        if not supplier:
            messages.error(
                request,
                "O XML não traz um CNPJ de fornecedor válido — não é possível importar automaticamente. "
                "Cadastre o fornecedor manualmente e lance a nota pela opção \"Lançar Manualmente\"."
            )
            return redirect('purchase_invoice_import_preview')

        duplicate_invoice = PurchaseInvoice.find_duplicate(
            supplier=supplier,
            number=parsed['number'],
            series=parsed['series'],
            access_key=parsed['access_key'],
        )
        if duplicate_invoice:
            messages.error(
                request,
                f"Esta nota já foi importada (registro #{duplicate_invoice.pk}, "
                f"status: {duplicate_invoice.get_status_display()})."
            )
            return redirect('purchase_invoice_detail', pk=duplicate_invoice.pk)

        matched_items, _unmatched = match_products(parsed['items'])
        created_products = create_missing_products(matched_items, supplier=supplier)

        # Modo Bling: opt-in checkbox pra atualizar campos fiscais de produtos
        # já cadastrados (só preenche campos vazios — nunca sobrescreve
        # configuração manual).
        regime_mismatch_labels = []
        updated_products = []
        if request.POST.get('update_existing_products'):
            for item in matched_items:
                product = item.get('product')
                if product is None or product in created_products:
                    continue
                changed = _update_product_fiscal(product, item)
                if changed:
                    updated_products.append(product.name)
            if updated_products:
                messages.info(
                    request,
                    f"Dados fiscais de {len(updated_products)} produto(s) existente(s) "
                    f"atualizados: {', '.join(updated_products[:5])}."
                )

        # Mark itens com CST/CSOSN divergente do regime da empresa — para o
        # preview renderizado abaixo (torna visível que ficou em branco).
        from fiscal.models import NFeConfig
        try:
            cfg = NFeConfig.load()
            is_simples = cfg.regime_tributario in (1, 2)
        except Exception:
            is_simples = None
        for item in matched_items:
            icms_raw = item.get('icms_raw') or ''
            if icms_raw and ((len(icms_raw) == 2 and is_simples) or (len(icms_raw) == 3 and not is_simples)):
                item['regime_mismatch'] = True
                regime_mismatch_labels.append(item['name'])
        if regime_mismatch_labels:
            messages.warning(
                request,
                f"{len(regime_mismatch_labels)} item(ns) tinham CST/CSOSN divergente do regime da empresa "
                f"e ficaram sem CST ICMS/CSOSN no cadastro. Revise manualmente: "
                f"{', '.join(regime_mismatch_labels[:5])}."
            )

        with transaction.atomic():
            locked_supplier = Supplier.objects.select_for_update().get(pk=supplier.pk)
            duplicate_invoice = PurchaseInvoice.find_duplicate(
                supplier=locked_supplier,
                number=parsed['number'],
                series=parsed['series'],
                access_key=parsed['access_key'],
            )
            if not duplicate_invoice:
                invoice_kwargs = dict(
                    supplier=locked_supplier,
                    number=parsed['number'],
                    series=parsed['series'],
                    access_key=''.join(filter(str.isdigit, parsed['access_key'] or '')),
                    generate_expense=bool(request.POST.get('generate_expense')),
                    expense_due_date=request.POST.get('expense_due_date') or None,
                    status=PurchaseInvoice.Status.RASCUNHO,
                )
                if parsed['issue_date']:
                    invoice_kwargs['issue_date'] = parsed['issue_date']
                invoice = PurchaseInvoice.objects.create(**invoice_kwargs)

                for item in matched_items:
                    PurchaseInvoiceItem.objects.create(
                        invoice=invoice,
                        product=item['product'],
                        quantity=item['quantity'],
                        unit_cost=item['unit_cost'],
                    )

        if duplicate_invoice:
            messages.error(
                request,
                f"Esta nota já foi importada (registro #{duplicate_invoice.pk}, "
                f"status: {duplicate_invoice.get_status_display()})."
            )
            return redirect('purchase_invoice_detail', pk=duplicate_invoice.pk)

        if supplier_created:
            messages.info(request, f"Fornecedor \"{supplier.display_name}\" cadastrado automaticamente a partir do XML.")
        if created_products:
            names = ", ".join(p.name for p in created_products[:5])
            more = f" e mais {len(created_products) - 5}" if len(created_products) > 5 else ""
            messages.info(request, f"{len(created_products)} produto(s) cadastrados automaticamente: {names}{more}. Revise categoria e preço de venda depois.")
        messages.success(request, "Nota de entrada importada como rascunho. Confira os itens e clique em \"Lançar Nota\" pra dar entrada no estoque.")
        return redirect('purchase_invoice_detail', pk=invoice.pk)

    if request.method == 'POST':
        upload_form = PurchaseInvoiceXMLUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            xml_file = upload_form.cleaned_data['xml_file']
            filename = xml_file.name
            try:
                raw_bytes = xml_file.read()
                xml_file.seek(0)
                raw_preview = raw_bytes[:800].decode('utf-8', errors='replace')
            except Exception:
                raw_preview = ''

            try:
                parsed = parse_nfe_xml(xml_file)
            except NFeXMLParseError as e:
                xml_import_error = str(e)
        else:
            xml_import_error = "Escolha um arquivo XML antes de enviar."

        if parsed:
            matched_items, _unmatched = match_products(parsed['items'])
            supplier = find_supplier_by_cnpj(parsed['supplier_cnpj'])
            duplicate_invoice = PurchaseInvoice.find_duplicate(
                supplier=supplier,
                number=parsed['number'],
                series=parsed['series'],
                access_key=parsed['access_key'],
            )
            for item in matched_items:
                item['subtotal'] = item['quantity'] * item['unit_cost']
            total = sum((item['subtotal'] for item in matched_items), Decimal('0'))
    else:
        upload_form = PurchaseInvoiceXMLUploadForm()

    return render(request, 'services/purchase_invoices/purchase_invoice_import_preview.html', {
        'upload_form': upload_form,
        'parsed': parsed,
        'matched_items': matched_items,
        'supplier': supplier,
        'total': total,
        'xml_import_error': xml_import_error,
        'raw_preview': raw_preview,
        'filename': filename,
        'duplicate_invoice': duplicate_invoice,
    })


def _items_initial_from_post(post_data):
    """Reconstrói a lista de itens a partir do POST sem passar pelo formset
    ligado (bound) — evita que o simples fato de renderizar o template
    (que acessa `formset.errors`) dispare validação e mostre erros de campo
    obrigatório quando só queremos reexibir o que o usuário já tinha
    preenchido (ex.: reenvio do Importar sem escolher o arquivo de novo, ou
    XML que falhou ao ser lido)."""
    try:
        total = int(post_data.get('items-TOTAL_FORMS', 0))
    except (TypeError, ValueError):
        total = 0

    rows = []
    for i in range(total):
        if post_data.get(f'items-{i}-DELETE'):
            continue
        rows.append({
            'product': post_data.get(f'items-{i}-product') or None,
            'quantity': post_data.get(f'items-{i}-quantity') or None,
            'unit_cost': post_data.get(f'items-{i}-unit_cost') or None,
        })
    return rows


def _header_initial_from_post(post_data):
    return {
        'supplier': post_data.get('supplier') or None,
        'number': post_data.get('number', ''),
        'series': post_data.get('series', ''),
        'access_key': post_data.get('access_key', ''),
        'issue_date': post_data.get('issue_date') or None,
        'notes': post_data.get('notes', ''),
        'generate_expense': bool(post_data.get('generate_expense')),
        'expense_due_date': post_data.get('expense_due_date') or None,
    }


def _xml_item_formset_class(extra):
    """FormSet com `extra` dinâmico — usado só para pré-popular as linhas de
    itens extraídas do XML antes de salvar (o formset padrão da tela manual
    usa extra=0 e adiciona linhas via JS)."""
    return inlineformset_factory(
        PurchaseInvoice, PurchaseInvoiceItem,
        form=PurchaseInvoiceItemForm,
        extra=extra, can_delete=True, min_num=1, validate_min=True,
    )


class PurchaseInvoiceListView(LoginRequiredMixin, ManagerRequiredMixin, ListView):
    model = PurchaseInvoice
    template_name = 'services/purchase_invoices/purchase_invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20

    def get_queryset(self):
        queryset = PurchaseInvoice.objects.select_related('supplier').order_by('-issue_date', '-created_at')
        status = self.request.GET.get('status')
        supplier_id = self.request.GET.get('supplier')
        if status:
            queryset = queryset.filter(status=status)
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = PurchaseInvoice.Status.choices
        context['suppliers'] = Supplier.objects.order_by('name')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_supplier'] = self.request.GET.get('supplier', '')
        return context


@login_required
@user_passes_test(is_manager)
def purchase_invoice_create(request):
    header_initial = None
    items_initial = None
    xml_import_error = ''

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'import_xml':
            upload_form = PurchaseInvoiceXMLUploadForm(request.POST, request.FILES)
            parsed = None

            if upload_form.is_valid():
                try:
                    parsed = parse_nfe_xml(upload_form.cleaned_data['xml_file'])
                except NFeXMLParseError as e:
                    xml_import_error = str(e)
                    messages.error(request, f"Não foi possível ler o XML: {e}")
            else:
                # O navegador limpa o campo de arquivo a cada envio do formulário —
                # se o usuário reenviar sem escolher o XML de novo, não há nada pra
                # importar. Mantém tudo que já estava preenchido em vez de zerar.
                xml_import_error = "Nenhum arquivo XML foi enviado. Escolha o arquivo antes de clicar em Importar."
                messages.error(request, xml_import_error)

            if parsed:
                matched_items, _unmatched = match_products(parsed['items'])

                supplier, supplier_created = get_or_create_supplier_from_xml(parsed)
                created_products = create_missing_products(matched_items, supplier=supplier)

                header_initial = {
                    'supplier': supplier.id if supplier else None,
                    'number': parsed['number'],
                    'series': parsed['series'],
                    'access_key': parsed['access_key'],
                    'issue_date': parsed['issue_date'] or None,
                }
                items_initial = [
                    {
                        'product': item['product'].id if item['product'] else None,
                        'quantity': item['quantity'],
                        'unit_cost': item['unit_cost'],
                    }
                    for item in matched_items
                ]

                if supplier_created:
                    messages.info(
                        request,
                        f"Fornecedor \"{supplier.display_name}\" (CNPJ {parsed['supplier_cnpj']}) não existia e foi "
                        "cadastrado automaticamente com os dados do XML. Revise o cadastro se precisar completar algo."
                    )
                elif not supplier:
                    messages.warning(
                        request,
                        f"O XML não traz um CNPJ de fornecedor válido ({parsed['supplier_cnpj'] or 's/ CNPJ'}). "
                        "Selecione o fornecedor manualmente."
                    )

                if created_products:
                    names = ", ".join(p.name for p in created_products[:5])
                    more = f" e mais {len(created_products) - 5}" if len(created_products) > 5 else ""
                    messages.info(
                        request,
                        f"{len(created_products)} produto(s) não existiam no catálogo e foram cadastrados "
                        f"automaticamente a partir do XML: {names}{more}. Revise categoria, preço de venda e "
                        "estoque mínimo depois — o XML só traz nome, código/EAN, unidade, NCM/CFOP e custo."
                    )

                messages.success(request, f"XML importado: {len(items_initial)} item(ns) prontos. Confira e salve.")

                form = PurchaseInvoiceForm(initial=header_initial)
                formset = _xml_item_formset_class(extra=max(len(items_initial or []), 1))(initial=items_initial)
            else:
                # Sem XML novo pra processar (arquivo não enviado ou XML não
                # reconhecido): reconstrói o formulário a partir do que veio no
                # POST sem "ligar" (bind) o form/formset a ele — só usando
                # `initial`. Isso preserva o que o usuário já tinha preenchido
                # sem disparar validação (o simples fato de renderizar
                # `{{ formset.errors }}" no template já validaria um formset
                # bound, mostrando "campo obrigatório" pra linhas que o
                # usuário nem tentou salvar ainda).
                header_initial = _header_initial_from_post(request.POST)
                rows = _items_initial_from_post(request.POST)
                form = PurchaseInvoiceForm(initial=header_initial)
                formset = _xml_item_formset_class(extra=max(len(rows), 1))(initial=rows)

        elif action in ('save_draft', 'save_and_launch'):
            form = PurchaseInvoiceForm(request.POST)
            formset = PurchaseInvoiceItemFormSet(request.POST)

            if form.is_valid() and formset.is_valid():
                with transaction.atomic():
                    supplier = Supplier.objects.select_for_update().get(pk=form.cleaned_data['supplier'].pk)
                    duplicate = PurchaseInvoice.find_duplicate(
                        supplier=supplier,
                        number=form.cleaned_data.get('number'),
                        series=form.cleaned_data.get('series'),
                        access_key=form.cleaned_data.get('access_key'),
                    )
                    if duplicate:
                        form.add_error(
                            None,
                            f'Esta nota já foi cadastrada (registro #{duplicate.pk}, '
                            f'status: {duplicate.get_status_display()}).'
                        )
                    else:
                        invoice = form.save(commit=False)
                        invoice.status = PurchaseInvoice.Status.RASCUNHO
                        invoice.save()
                        formset.instance = invoice
                        formset.save()

                if duplicate:
                    return render(request, 'services/purchase_invoices/purchase_invoice_form.html', {
                        'form': form, 'formset': formset,
                        'upload_form': PurchaseInvoiceXMLUploadForm(),
                        'title': 'Nova Nota de Entrada',
                    })

                if action == 'save_and_launch':
                    try:
                        invoice.launch(user=request.user)
                        messages.success(request, "Nota de entrada lançada com sucesso — estoque e custo dos produtos atualizados.")
                    except ValueError as e:
                        messages.error(request, f"Nota salva como rascunho, mas não pôde ser lançada: {e}")
                else:
                    messages.success(request, "Nota de entrada salva como rascunho.")
                return redirect('purchase_invoice_detail', pk=invoice.pk)
        else:
            form = PurchaseInvoiceForm()
            formset = PurchaseInvoiceItemFormSet()
    else:
        form = PurchaseInvoiceForm()
        formset = PurchaseInvoiceItemFormSet()

    return render(request, 'services/purchase_invoices/purchase_invoice_form.html', {
        'form': form,
        'formset': formset,
        'upload_form': PurchaseInvoiceXMLUploadForm(),
        'title': 'Nova Nota de Entrada',
        'xml_import_error': xml_import_error,
    })


class PurchaseInvoiceDetailView(LoginRequiredMixin, ManagerRequiredMixin, DetailView):
    model = PurchaseInvoice
    template_name = 'services/purchase_invoices/purchase_invoice_detail.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.select_related('product').all()
        return context


@login_required
@user_passes_test(is_manager)
def purchase_invoice_launch(request, pk):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    if request.method == 'POST':
        try:
            invoice.launch(user=request.user)
            messages.success(request, "Nota de entrada lançada com sucesso — estoque e custo dos produtos atualizados.")
        except ValueError as e:
            messages.error(request, str(e))
    return redirect('purchase_invoice_detail', pk=invoice.pk)


@login_required
@user_passes_test(is_manager)
def purchase_invoice_cancel(request, pk):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    if request.method == 'POST':
        try:
            invoice.cancel()
            messages.success(request, "Nota de entrada cancelada.")
        except ValueError as e:
            messages.error(request, str(e))
    return redirect('purchase_invoice_detail', pk=invoice.pk)
