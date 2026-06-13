from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Q, F, Avg, Count
from django.db import transaction
from django.utils import timezone
from django.core.paginator import Paginator
from decimal import Decimal
from datetime import datetime, date
import json
import calendar as _cal
from .models import ServiceOrderTask, ServiceOrderTeam, Professional, User, ServicePayment, Sale, SaleItem, Product, StockMovement, PaymentMethod, SalePayment, Expense, ExpenseInstallment, Client, RecurrenceRule, FinanceSettings, BankAccount, InstallmentPayment, PaymentAttachment, FinancialCategory, SaleReturn, SaleReturnItem, ProductVariant, SaleSettings, ServiceOrder, Billing, Installment
from .forms import SaleForm, SaleItemFormSet, PaymentMethodForm, ExpenseForm, ExpenseInstallmentFormSet, FinanceSettingsForm, SaleSettingsForm
from .expense_engine import generate_expense_installments
from django.http import HttpResponse, JsonResponse
import csv
import logging
from io import BytesIO

logger = logging.getLogger(__name__)
from dateutil.relativedelta import relativedelta
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Billing, Installment

def is_manager(user):
    return user.is_superuser or user.role in [User.Roles.ADMIN, User.Roles.MANAGER]

@login_required
@user_passes_test(is_manager)
def payment_method_list(request):
    methods = PaymentMethod.objects.all().order_by('descricao')
    context = {
        'methods': methods,
        'title': 'Métodos de Pagamento',
    }
    return render(request, 'services/finance/payment_method_list.html', context)

@login_required
@user_passes_test(is_manager)
def payment_method_create(request):
    if request.method == 'POST':
        form = PaymentMethodForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Método de pagamento criado com sucesso!")
            return redirect('payment_method_list')
    else:
        form = PaymentMethodForm()
    
    context = {
        'form': form,
        'title': 'Novo Método de Pagamento',
        'is_edit': False
    }
    return render(request, 'services/finance/payment_method_form.html', context)

@login_required
@user_passes_test(is_manager)
def payment_method_edit(request, pk):
    method = get_object_or_404(PaymentMethod, pk=pk)
    if request.method == 'POST':
        form = PaymentMethodForm(request.POST, instance=method)
        if form.is_valid():
            form.save()
            messages.success(request, "Método de pagamento atualizado com sucesso!")
            return redirect('payment_method_list')
    else:
        form = PaymentMethodForm(instance=method)
    
    context = {
        'form': form,
        'title': f'Editar: {method.descricao}',
        'is_edit': True
    }
    return render(request, 'services/finance/payment_method_form.html', context)

@login_required
@user_passes_test(is_manager)
def payment_method_toggle(request, pk):
    method = get_object_or_404(PaymentMethod, pk=pk)
    method.ativo = not method.ativo
    method.save()
    status = "ativado" if method.ativo else "desativado"
    messages.success(request, f"Método de pagamento {method.descricao} {status} com sucesso!")
    return redirect('payment_method_list')

@login_required
@user_passes_test(is_manager)
def finance_dashboard(request):
    now = timezone.now()
    month = int(request.GET.get('month', now.month))
    year = int(request.GET.get('year', now.year))

    tz = timezone.get_current_timezone()
    start_of_month = timezone.make_aware(datetime(year, month, 1), tz)
    if month == 12:
        end_of_month = timezone.make_aware(datetime(year + 1, 1, 1), tz)
    else:
        end_of_month = timezone.make_aware(datetime(year, month + 1, 1), tz)

    try:
        from dateutil.relativedelta import relativedelta as _rdelta
        _has_rdelta = True
    except ImportError:
        _has_rdelta = False

    def _prev_month(y, m, n):
        if _has_rdelta:
            d = date(y, m, 1) - _rdelta(months=n)
            return d.year, d.month
        total = m - 1 - n
        return y + total // 12, total % 12 + 1

    bi_months_meta = []
    for i in range(5, -1, -1):
        my, mm = _prev_month(year, month, i)
        _, ld = _cal.monthrange(my, mm)
        ms = date(my, mm, 1)
        me = date(my, mm, ld)
        tz_obj = timezone.get_current_timezone()
        ms_dt = timezone.make_aware(datetime(my, mm, 1), tz_obj)
        me_dt = timezone.make_aware(datetime(my, mm, ld, 23, 59, 59), tz_obj)
        label = f"{mm:02d}/{my}"
        bi_months_meta.append((my, mm, ms, me, ms_dt, me_dt, label))

    monthly_revenue_labels = [m[6] for m in bi_months_meta]
    monthly_revenue_values = []
    for my, mm, ms, me, ms_dt, me_dt, _ in bi_months_meta:
        rev = SalePayment.objects.filter(
            data_pagamento__date__gte=ms,
            data_pagamento__date__lte=me,
        ).aggregate(total=Sum('valor_bruto'))['total'] or Decimal('0')
        monthly_revenue_values.append(float(rev))

    monthly_expense_values = []
    for my, mm, ms, me, ms_dt, me_dt, _ in bi_months_meta:
        exp = InstallmentPayment.objects.filter(
            payment_date__gte=ms,
            payment_date__lte=me,
            is_discount=False,
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        monthly_expense_values.append(float(exp))

    _cur_my, _cur_mm, _cur_ms, _cur_me = bi_months_meta[-1][:4]
    dre_receita = SalePayment.objects.filter(
        data_pagamento__date__gte=_cur_ms,
        data_pagamento__date__lte=_cur_me,
    ).aggregate(total=Sum('valor_bruto'))['total'] or Decimal('0')

    dre_despesas = InstallmentPayment.objects.filter(
        payment_date__gte=_cur_ms,
        payment_date__lte=_cur_me,
        is_discount=False,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    dre_resultado = dre_receita - dre_despesas

    avg_ticket = ServiceOrder.objects.filter(
        status__in=[ServiceOrder.Status.FINALIZADO, ServiceOrder.Status.CONCLUIDA],
        finished_at__gte=bi_months_meta[-1][4],
        finished_at__lte=bi_months_meta[-1][5],
        estimated_value__isnull=False,
    ).aggregate(avg=Avg('estimated_value'))['avg'] or Decimal('0')

    today_date = date.today()
    total_active_installments = Installment.objects.exclude(
        status__in=[Installment.Status.PAGO, Installment.Status.CANCELADO]
    ).count()
    overdue_installments = Installment.objects.filter(
        Q(status=Installment.Status.ATRASADO) |
        Q(status=Installment.Status.PENDENTE, due_date__lt=today_date)
    ).count()
    inadimplencia_rate = round(
        (overdue_installments / total_active_installments * 100)
        if total_active_installments > 0 else 0, 1
    )

    top_clients = (
        Billing.objects
        .exclude(status=Billing.Status.CANCELADO)
        .values(cliente=F('client__name'))
        .annotate(total=Sum('total_amount'))
        .order_by('-total')[:10]
    )
    top_clients_labels = json.dumps([c['cliente'] or 'Sem nome' for c in top_clients])
    top_clients_values = json.dumps([float(c['total']) for c in top_clients])

    os_count = ServiceOrder.objects.filter(
        status__in=[ServiceOrder.Status.FINALIZADO, ServiceOrder.Status.CONCLUIDA],
        finished_at__gte=start_of_month,
        finished_at__lt=end_of_month,
    ).count()

    context = {
        'months': range(1, 13),
        'years': range(now.year - 2, now.year + 1),
        'selected_month': month,
        'selected_year': year,
        'monthly_revenue_labels_json': json.dumps(monthly_revenue_labels),
        'monthly_revenue_json': json.dumps(monthly_revenue_values),
        'monthly_expense_json': json.dumps(monthly_expense_values),
        'dre_receita': dre_receita,
        'dre_despesas': dre_despesas,
        'dre_resultado': dre_resultado,
        'avg_ticket': avg_ticket,
        'overdue_installments': overdue_installments,
        'inadimplencia_rate': inadimplencia_rate,
        'top_clients_labels_json': top_clients_labels,
        'top_clients_values_json': top_clients_values,
        'os_count': os_count,
    }

    return render(request, 'services/finance/dashboard.html', context)


@login_required
@user_passes_test(is_manager)
def finance_commissions(request):
    now = timezone.now()
    month = int(request.GET.get('month', now.month))
    year = int(request.GET.get('year', now.year))
    professional_id = request.GET.get('professional')

    tz = timezone.get_current_timezone()
    start_of_month = timezone.make_aware(datetime(year, month, 1), tz)
    if month == 12:
        end_of_month = timezone.make_aware(datetime(year + 1, 1, 1), tz)
    else:
        end_of_month = timezone.make_aware(datetime(year, month + 1, 1), tz)

    finance_settings, _ = FinanceSettings.objects.get_or_create(pk=1)
    commission_enabled = finance_settings.enable_commission

    billable_statuses = [
        ServiceOrderTask.TaskStatus.CONCLUIDO,
        ServiceOrderTask.TaskStatus.PARCIALMENTE_EXECUTADO,
    ]
    allocations = ServiceOrderTeam.objects.filter(
        task__task_type=ServiceOrderTask.TaskType.EXECUCAO,
        task__status__in=billable_statuses,
        task__finished_at__gte=start_of_month,
        task__finished_at__lt=end_of_month
    ).select_related('task__service_order', 'professional', 'role').order_by('professional__name', 'task__finished_at')

    if professional_id:
        allocations = allocations.filter(professional_id=professional_id)

    grouped_data = {}
    total_commission_global = Decimal('0')
    total_services_value_global = Decimal('0')
    total_base_salary_global = Decimal('0')
    professionals_seen = set()

    for alloc in allocations:
        p_id = str(alloc.professional.id)
        if p_id not in grouped_data:
            grouped_data[p_id] = {
                'professional_name': alloc.professional.name,
                'base_salary': alloc.professional.base_salary or Decimal('0'),
                'items': [],
                'total_commission': Decimal('0'),
                'total_services_value': Decimal('0'),
            }

        task_value = alloc.task.billing_value or Decimal('0')
        comm_rate = alloc.role.commission_rate if alloc.role else Decimal('0')
        commission_value = (task_value * (comm_rate / 100))

        item = {
            'date': alloc.task.finished_at,
            'order_id': alloc.task.service_order.id,
            'os_number': alloc.task.service_order.number,
            'client_name': alloc.task.service_order.client_property.client.name,
            'address': alloc.task.service_order.client_property.address,
            'role': alloc.role.name if alloc.role else '---',
            'task_value': task_value,
            'comm_rate': comm_rate,
            'commission_value': commission_value,
        }

        grouped_data[p_id]['items'].append(item)
        grouped_data[p_id]['total_commission'] += commission_value
        grouped_data[p_id]['total_services_value'] += task_value

        total_commission_global += commission_value
        total_services_value_global += task_value

        if alloc.professional.id not in professionals_seen:
            professionals_seen.add(alloc.professional.id)
            total_base_salary_global += alloc.professional.base_salary or Decimal('0')

    report_items = []
    for p_id, data in grouped_data.items():
        for item in data['items']:
            item_with_name = item.copy()
            item_with_name['professional'] = data['professional_name']
            report_items.append(item_with_name)

    if request.GET.get('export') == 'xlsx':
        return export_finance_xlsx(grouped_data, month, year)
    elif request.GET.get('export') == 'pdf':
        return export_finance_pdf(grouped_data, month, year)

    context = {
        'report_items': report_items,
        'grouped_data': grouped_data,
        'commission_enabled': commission_enabled,
        'total_services_count': len(report_items),
        'total_services_value': total_services_value_global,
        'total_commission': total_commission_global,
        'total_base_salary': total_base_salary_global,
        'total_with_salary': total_commission_global + total_base_salary_global,
        'months': range(1, 13),
        'years': range(now.year - 2, now.year + 1),
        'selected_month': month,
        'selected_year': year,
        'selected_professional': professional_id,
        'professionals': Professional.objects.filter(is_active=True),
    }

    return render(request, 'services/finance/commissions.html', context)


def export_finance_xlsx(grouped_data, month, year):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Financeiro"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    group_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    for p_id, data in grouped_data.items():
        # Professional Header
        ws.append([f"Colaborador: {data['professional_name']}"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=8)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=ws.max_row, column=1).fill = group_fill

        # Table Headers
        headers = ['Data', 'OS', 'Cliente', 'Endereço', 'Função', 'Vlr. OS (Est.)', 'Comissão (%)', 'Vlr. Comissão']
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Data Rows
        for item in data['items']:
            ws.append([
                item['date'].strftime('%d/%m/%Y'),
                item['os_number'],
                item['client_name'],
                item['address'],
                item['role'],
                float(item['task_value']),
                float(item['comm_rate']),
                float(item['commission_value'])
            ])
            curr_row = ws.max_row
            ws.cell(row=curr_row, column=6).number_format = '"R$ "#,##0.00'
            ws.cell(row=curr_row, column=8).number_format = '"R$ "#,##0.00'
            ws.cell(row=curr_row, column=7).number_format = '0.0"%"'

        # Subtotals
        ws.append(['', '', '', '', '', 'Total Comissão:', '', float(data['total_commission'])])
        ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=8).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=8).number_format = '"R$ "#,##0.00'
        
        ws.append(['', '', '', '', '', 'Salário Base:', '', float(data['base_salary'])])
        ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=8).number_format = '"R$ "#,##0.00'
        
        ws.append(['', '', '', '', '', 'TOTAL GERAL:', '', float(data['total_commission'] + data['base_salary'])])
        last_row = ws.max_row
        ws.cell(row=last_row, column=6).font = Font(bold=True)
        ws.cell(row=last_row, column=8).font = Font(bold=True, color="FF0000")
        ws.cell(row=last_row, column=8).number_format = '"R$ "#,##0.00'
        
        ws.append([]) # Spacer
        ws.append([]) # Spacer

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="servicos_por_colaborador_{month}_{year}.xlsx"'
    wb.save(response)
    return response

def export_finance_pdf(grouped_data, month, year):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    for p_id, data in grouped_data.items():
        # Header for each professional
        elements.append(Paragraph("Serviços por colaborador", styles['Title']))
        now = timezone.now().astimezone(timezone.get_current_timezone())
        timestamp = now.strftime('%d/%m/%Y %H:%M:%S')
        elements.append(Paragraph(f"Colaborador: <b>{data['professional_name']}</b> | Período: {month:02d}/{year} | Gerado em: {timestamp}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table
        table_data = [['Data', 'OS', 'Cliente', 'Endereço', 'Função', 'Vlr. Est. OS', 'Comis %', 'Vlr. Comis.']]
        for item in data['items']:
            table_data.append([
                item['date'].strftime('%d/%m/%Y'),
                str(item['os_number']),
                Paragraph(item['client_name'], styles['Normal']),
                Paragraph(item['address'], styles['Normal']),
                item['role'],
                f"R$ {item['task_value']:,.2f}",
                f"{item['comm_rate']}%",
                f"R$ {item['commission_value']:,.2f}"
            ])
        
        table = Table(table_data, repeatRows=1, colWidths=[60, 40, 140, 180, 100, 80, 60, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 15))
        
        # Professional Totals
        summary_data = [
            ['Total Comissão', f"R$ {data['total_commission']:,.2f}"],
            ['Salário Base', f"R$ {data['base_salary']:,.2f}"],
            ['TOTAL GERAL', f"R$ {(data['total_commission'] + data['base_salary']):,.2f}"]
        ]
        summary_table = Table(summary_data, colWidths=[150, 100])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.whitesmoke),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (1, 2), (1, 2), colors.red),
        ]))
        elements.append(summary_table)
        
        # Signature
        elements.append(Spacer(1, 40))
        sig_data = [
            ['__________________________________________', '__________________________________________'],
            [f'Assinatura: {data["professional_name"]}', 'Data']
        ]
        sig_table = Table(sig_data, colWidths=[300, 300])
        sig_table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTSIZE', (0, 1), (-1, 1), 8)]))
        elements.append(sig_table)
        
        # Page break for next professional
        from reportlab.platypus import PageBreak
        elements.append(PageBreak())
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="servicos_por_colaborador_{month}_{year}.pdf"'
    response.write(pdf)
    return response

@login_required
def finance_professional_payments(request):
    """
    Lista pagamentos recebidos por técnicos.
    Se for gerente, vê todos (com filtro).
    Se for técnico, vê apenas os seus.
    """
    professional_id = request.GET.get('professional')
    status_filter = request.GET.get('status', 'PENDENTE')
    
    payments = ServicePayment.objects.select_related(
        'order', 'order__client_property__client', 'received_by'
    ).order_by('-paid_at')
    
    # Filtro de status
    if status_filter in ['PENDENTE', 'CONFIRMED']:
        payments = payments.filter(status=status_filter)
    
    # Filtro de permissão/profissional
    if not is_manager(request.user):
        try:
            professional = request.user.professional_profile
            payments = payments.filter(received_by=professional)
            professionals = [professional]
            selected_professional = str(professional.id)
        except Professional.DoesNotExist:
            payments = ServicePayment.objects.none()
            professionals = []
            selected_professional = None
    else:
        professionals = Professional.objects.filter(is_active=True)
        if professional_id:
            payments = payments.filter(received_by_id=professional_id)
            selected_professional = professional_id
        else:
            selected_professional = None
        
    paginator = Paginator(payments, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'payments': page_obj,  # Antigamente era 'payments': payments
        'page_obj': page_obj,
        'professionals': professionals,
        'selected_professional': selected_professional,
        'status_filter': status_filter,
        'title': 'Gestão de Recebimentos',
        'is_manager': is_manager(request.user),
        'layout_base': 'base.html' if is_manager(request.user) else 'base_equipe.html'
    }
    return render(request, 'services/finance/professional_payments.html', context)

@login_required
@user_passes_test(is_manager)
def finance_confirm_payment(request, payment_id):
    """Dá baixa em um pagamento individual."""
    payment = get_object_or_404(ServicePayment, id=payment_id)
    
    if request.method == 'POST':
        payment.status = 'CONFIRMED'
        payment.confirmed_at = timezone.now()
        payment.confirmed_by = request.user
        payment.save()
        messages.success(request, f'Pagamento de R$ {payment.amount} (OS #{payment.order.number}) baixado com sucesso.')
        
    return redirect('finance_professional_payments')

@login_required
@user_passes_test(is_manager)
def finance_bulk_confirm_payments(request):
    """Dá baixa em múltiplos pagamentos selecionados."""
    if request.method == 'POST':
        payment_ids = request.POST.getlist('payment_ids')
        if payment_ids:
            # We iterate to ensure order.update_status() is called via save() or explicitly
            payments = ServicePayment.objects.filter(id__in=payment_ids, status='PENDENTE')
            count = 0
            for p in payments:
                p.status = 'CONFIRMED'
                p.confirmed_at = timezone.now()
                p.confirmed_by = request.user
                p.save()
                count += 1
                
            messages.success(request, f'{count} pagamentos foram baixados com sucesso.')
        else:
            messages.warning(request, 'Nenhum pagamento selecionado.')
            
    return redirect('finance_professional_payments')


# --- MÓDULO DE VENDAS (PDV) ---

@login_required
@user_passes_test(is_manager)
def sale_list(request):
    sales = Sale.objects.all().select_related('client', 'user').order_by('-created_at')

    q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    seller_id = request.GET.get('seller', '').strip()

    if q:
        sales = sales.filter(
            Q(number__icontains=q) |
            Q(client__name__icontains=q) |
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(external_po_number__icontains=q)
        )

    if status_filter:
        sales = sales.filter(status=status_filter)

    if date_from:
        try:
            sales = sales.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass

    if date_to:
        try:
            sales = sales.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    if seller_id:
        sales = sales.filter(user_id=seller_id)

    finalized_statuses = [Sale.Status.FINALIZADA, Sale.Status.ATENDIDO, Sale.Status.RECEBIDO]
    kpi_qs = sales.filter(status__in=finalized_statuses)
    total_vendido = kpi_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    count_vendas = kpi_qs.count()
    ticket_medio = (total_vendido / count_vendas) if count_vendas > 0 else Decimal('0.00')
    count_rascunhos = sales.filter(status=Sale.Status.RASCUNHO).count()

    sellers = User.objects.filter(id__in=Sale.objects.values_list('user_id', flat=True)).distinct()

    paginator = Paginator(sales, 20)
    sales_page = paginator.get_page(request.GET.get('page'))

    context = {
        'sales': sales_page,
        'title': 'Histórico de Vendas',
        'total_vendido': total_vendido,
        'count_vendas': count_vendas,
        'ticket_medio': ticket_medio,
        'count_rascunhos': count_rascunhos,
        'sellers': sellers,
        'all_statuses': Sale.Status.choices,
        'q': q,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'seller_id': seller_id,
    }
    return render(request, 'services/sale_list.html', context)

@login_required
@user_passes_test(is_manager)
def sale_create(request):
    if request.method == 'POST':
        form = SaleForm(request.POST)
        formset = SaleItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    is_draft = request.POST.get('save_as_draft') == '1'

                    sale = form.save(commit=False)
                    sale.user = request.user
                    sale.save()

                    formset.instance = sale
                    items = formset.save()

                    total = Decimal('0.00')
                    for item in items:
                        total += item.subtotal
                        if not is_draft:
                            target = item.variant if item.variant else item.product
                            target.reduce_stock(
                                item.quantity,
                                user=request.user,
                                reason=StockMovement.Reason.VENDA_DIRETA,
                                notes=f"Venda Direta: #{sale.number}"
                            )

                    sale.total_amount = total - sale.discount + sale.surcharge
                    sale.status = Sale.Status.RASCUNHO if is_draft else Sale.Status.FINALIZADA
                    sale.stock_reduced = not is_draft
                    sale.save()

                    if not is_draft:
                        installment_dates = request.POST.getlist('installment_due_date[]')
                        installment_amounts = request.POST.getlist('installment_amount[]')
                        installment_methods = request.POST.getlist('installment_method_id[]')

                        if installment_dates:
                            installments_data = []
                            for d, a, m in zip(installment_dates, installment_amounts, installment_methods):
                                installments_data.append({
                                    'due_date': d,
                                    'amount': Decimal(a),
                                    'payment_method_id': m if m else None
                                })
                            from .utils.finance import create_billing_for_sale
                            create_billing_for_sale(sale, installments_data)

                    action = "salvo como rascunho" if is_draft else "realizada com sucesso"
                    messages.success(request, f"Venda #{sale.number} {action}!")
                    return redirect('sale_detail', number=sale.number)
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                messages.error(request, f"Erro ao processar venda: {str(e)}")
        else:
            messages.error(request, "Por favor, corrija os erros no formulário.")
    else:
        form = SaleForm()
        formset = SaleItemFormSet(queryset=SaleItem.objects.none())
    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Nova Venda (PDV)',
        'products': Product.objects.filter(is_active=True),
        'clients': Client.objects.all().order_by('name'),
        'payment_methods': PaymentMethod.objects.filter(ativo=True)
    }
    return render(request, 'services/sale_form.html', context)

@login_required
@user_passes_test(is_manager)
def sale_detail(request, number):
    sale = get_object_or_404(Sale, number=number)
    can_edit = sale.can_be_edited()
    
    if request.method == 'POST':
        if not can_edit:
            messages.error(request, "Esta venda não pode ser editada pois já foi finalizada ou possui parcelas no Contas a Receber.")
            return redirect('sale_detail', number=sale.number)
            
        form = SaleForm(request.POST, instance=sale)
        formset = SaleItemFormSet(request.POST, instance=sale)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    is_draft = request.POST.get('save_as_draft') == '1'

                    if sale.stock_reduced:
                        for item in sale.items.all():
                            target = item.variant if item.variant else item.product
                            target.increase_stock(
                                item.quantity,
                                user=request.user,
                                reason=StockMovement.Reason.DEVOLUCAO,
                                notes=f"Estorno para Reedição da Venda: #{sale.number}"
                            )

                    sale = form.save(commit=False)
                    sale.save()

                    formset.instance = sale
                    items = formset.save()

                    total = Decimal('0.00')
                    for item in sale.items.all():
                        total += item.subtotal
                        if not is_draft:
                            target = item.variant if item.variant else item.product
                            target.reduce_stock(
                                item.quantity,
                                user=request.user,
                                reason=StockMovement.Reason.VENDA_DIRETA,
                                notes=f"Venda Direta (Editada): #{sale.number}"
                            )

                    sale.total_amount = total - sale.discount + sale.surcharge
                    sale.status = Sale.Status.RASCUNHO if is_draft else Sale.Status.FINALIZADA
                    sale.stock_reduced = not is_draft
                    sale.save()

                    if not is_draft:
                        installment_dates = request.POST.getlist('installment_due_date[]')
                        installment_amounts = request.POST.getlist('installment_amount[]')
                        installment_methods = request.POST.getlist('installment_method_id[]')

                        if installment_dates:
                            installments_data = []
                            for d, a, m in zip(installment_dates, installment_amounts, installment_methods):
                                installments_data.append({
                                    'due_date': d,
                                    'amount': Decimal(a),
                                    'payment_method_id': m if m else None
                                })
                            from .utils.finance import create_billing_for_sale
                            create_billing_for_sale(sale, installments_data)

                    messages.success(request, f"Venda #{sale.number} atualizada com sucesso!")
                    return redirect('sale_detail', number=sale.number)
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                messages.error(request, f"Erro ao atualizar venda: {str(e)}")
        else:
            messages.error(request, "Por favor, corrija os erros no formulário.")
    else:
        form = SaleForm(instance=sale)
        formset = SaleItemFormSet(instance=sale)
    
    # Se já tiver billing, carregar as parcelas para o template
    initial_installments = []
    if hasattr(sale, 'billing'):
        for inst in sale.billing.installments.all():
            initial_installments.append({
                'due_date': inst.due_date.isoformat(),
                'amount': inst.amount,
                'method_id': inst.payment_method_id
            })

    context = {
        'sale': sale,
        'form': form,
        'formset': formset,
        'can_edit': can_edit,
        'is_detail': True,
        'title': f'Venda #{sale.number}',
        'products': Product.objects.filter(is_active=True),
        'clients': Client.objects.all().order_by('name'),
        'payment_methods': PaymentMethod.objects.filter(ativo=True),
        'initial_installments': initial_installments,
        'returns': sale.returns.all().prefetch_related('items__sale_item__product') if hasattr(sale, 'returns') else [],
    }
    return render(request, 'services/sale_form.html', context)

@login_required
@user_passes_test(is_manager)
def sale_cancel(request, number):
    sale = get_object_or_404(Sale, number=number)
    
    if sale.status == Sale.Status.CANCELADO:
        messages.warning(request, "Esta venda já está cancelada.")
        return redirect('sale_detail', number=sale.number)
        
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Reverter Estoque apenas se já tinha sido baixado
                if sale.stock_reduced:
                    for item in sale.items.all():
                        target = item.variant if item.variant else item.product
                        target.increase_stock(
                            item.quantity,
                            user=request.user,
                            reason=StockMovement.Reason.DEVOLUCAO,
                            notes=f"Estorno Venda Cancelada: #{sale.number}"
                        )

                sale.status = Sale.Status.CANCELADO
                sale.stock_reduced = False
                sale.save()

                # Cancelar Billing se existir
                if hasattr(sale, 'billing'):
                    sale.billing.status = Billing.Status.CANCELADO
                    sale.billing.save()
                    sale.billing.installments.update(status=Installment.Status.CANCELADO)

                messages.success(request, "Venda cancelada e estoque estornado com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro ao cancelar venda: {str(e)}")
            
    return redirect('sale_detail', number=sale.number)

@login_required
@user_passes_test(is_manager)
def sale_duplicate(request, number):
    original = get_object_or_404(Sale, number=number)
    with transaction.atomic():
        new_sale = Sale.objects.create(
            client=original.client,
            user=request.user,
            status=Sale.Status.RASCUNHO,
            discount=original.discount,
            surcharge=original.surcharge,
            indicador_presenca=original.indicador_presenca,
            modalidade_frete=original.modalidade_frete,
            forma_pagamento_sefaz=original.forma_pagamento_sefaz,
            notes_internal=original.notes_internal,
            notes_customer=original.notes_customer,
            payment_method=original.payment_method,
            commission_rate=original.commission_rate,
            external_po_number=original.external_po_number,
            delivery_name=original.delivery_name,
            delivery_cep=original.delivery_cep,
            delivery_address=original.delivery_address,
            delivery_number=original.delivery_number,
            delivery_complement=original.delivery_complement,
            delivery_neighborhood=original.delivery_neighborhood,
            delivery_city=original.delivery_city,
            delivery_state=original.delivery_state,
            stock_reduced=False,
        )
        for item in original.items.all():
            SaleItem.objects.create(
                sale=new_sale,
                product=item.product,
                variant=item.variant,
                quantity=item.quantity,
                unit_price=item.unit_price,
                discount=item.discount,
            )
        total = sum(i.subtotal for i in new_sale.items.all())
        new_sale.total_amount = total - new_sale.discount + new_sale.surcharge
        new_sale.save()

    messages.success(request, f"Venda #{original.number} duplicada como Rascunho #{new_sale.number}!")
    return redirect('sale_detail', number=new_sale.number)


@login_required
@user_passes_test(is_manager)
def api_product_stock(request, product_id):
    try:
        product = Product.objects.get(pk=product_id, is_active=True)
        variants = []
        if product.format == Product.Format.COM_VARIACOES:
            variants = list(product.variants.filter(is_active=True).values('id', 'name', 'current_stock', 'min_stock', 'additional_price'))
        return JsonResponse({
            'stock': float(product.current_stock),
            'min_stock': float(product.min_stock),
            'cost': float(product.preco_custo),
            'price': float(product.default_unit_price),
            'has_variants': product.format == Product.Format.COM_VARIACOES,
            'variants': variants,
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Produto não encontrado'}, status=404)


@login_required
@user_passes_test(is_manager)
def sale_export_csv(request):
    sales = Sale.objects.all().select_related('client', 'user').order_by('-created_at')

    q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    seller_id = request.GET.get('seller', '').strip()

    if q:
        sales = sales.filter(Q(number__icontains=q) | Q(client__name__icontains=q) | Q(user__username__icontains=q))
    if status_filter:
        sales = sales.filter(status=status_filter)
    if date_from:
        try:
            sales = sales.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            sales = sales.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    if seller_id:
        sales = sales.filter(user_id=seller_id)

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="vendas.csv"'
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['N°', 'Data', 'Cliente', 'Vendedor', 'Total', 'Desconto', 'Acréscimo', 'Status', 'PO Cliente', 'Prazo Entrega', 'Comissão %'])
    for sale in sales:
        writer.writerow([
            sale.number,
            sale.created_at.strftime('%d/%m/%Y %H:%M'),
            sale.client.display_name if sale.client else 'Consumidor Final',
            sale.user.get_full_name() or sale.user.username,
            str(sale.total_amount).replace('.', ','),
            str(sale.discount).replace('.', ','),
            str(sale.surcharge).replace('.', ','),
            sale.get_status_display(),
            sale.external_po_number or '',
            sale.delivery_date.strftime('%d/%m/%Y') if sale.delivery_date else '',
            str(sale.commission_rate).replace('.', ','),
        ])
    return response


@login_required
@user_passes_test(is_manager)
def sale_return_create(request, number):
    sale = get_object_or_404(Sale, number=number)

    if sale.status == Sale.Status.CANCELADO:
        messages.error(request, "Não é possível criar devolução para venda cancelada.")
        return redirect('sale_detail', number=number)

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        if not reason:
            messages.error(request, "Informe o motivo da devolução.")
            return redirect('sale_return_create', number=number)

        item_ids = request.POST.getlist('item_id[]')
        quantities = request.POST.getlist('quantity[]')
        restocks = request.POST.getlist('restock[]')

        items_to_return = []
        total_refund = Decimal('0.00')

        for item_id, qty_str in zip(item_ids, quantities):
            try:
                qty = Decimal(str(qty_str).replace(',', '.').strip())
                if qty <= 0:
                    continue
                sale_item = sale.items.get(pk=item_id)
                if qty > sale_item.quantity:
                    messages.error(request, f"Quantidade de devolução para {sale_item.product.name} excede a quantidade vendida.")
                    return redirect('sale_return_create', number=number)
                refund = (sale_item.unit_price - sale_item.discount / sale_item.quantity) * qty
                items_to_return.append((sale_item, qty, refund, str(item_id) in restocks))
                total_refund += refund
            except (ValueError, SaleItem.DoesNotExist):
                continue

        if not items_to_return:
            messages.error(request, "Selecione ao menos um item para devolver.")
            return redirect('sale_return_create', number=number)

        with transaction.atomic():
            sale_return = SaleReturn.objects.create(
                sale=sale,
                user=request.user,
                reason=reason,
                total_refund=total_refund,
                status=SaleReturn.Status.PENDENTE,
            )
            for sale_item, qty, refund, do_restock in items_to_return:
                SaleReturnItem.objects.create(
                    sale_return=sale_return,
                    sale_item=sale_item,
                    quantity=qty,
                    refund_amount=refund,
                    restock=do_restock,
                )

        messages.success(request, "Devolução registrada com sucesso! Aguardando aprovação.")
        return redirect('sale_detail', number=number)

    context = {
        'sale': sale,
        'title': f'Devolução — Venda #{sale.number}',
    }
    return render(request, 'services/sale_return_form.html', context)


@login_required
@user_passes_test(is_manager)
def sale_return_approve(request, return_id):
    sale_return = get_object_or_404(SaleReturn, pk=return_id)

    if sale_return.status != SaleReturn.Status.PENDENTE:
        messages.warning(request, "Esta devolução já foi processada.")
        return redirect('sale_detail', number=sale_return.sale.number)

    if request.method == 'POST':
        with transaction.atomic():
            for return_item in sale_return.items.all():
                if return_item.restock:
                    target = return_item.sale_item.variant if return_item.sale_item.variant else return_item.sale_item.product
                    target.increase_stock(
                        return_item.quantity,
                        user=request.user,
                        reason=StockMovement.Reason.DEVOLUCAO,
                        notes=f"Devolução #{sale_return.pk} — Venda #{sale_return.sale.number}"
                    )
            sale_return.status = SaleReturn.Status.APROVADA
            sale_return.save()
        messages.success(request, f"Devolução #{sale_return.pk} aprovada e estoque atualizado.")

    return redirect('sale_detail', number=sale_return.sale.number)


@login_required
@user_passes_test(is_manager)
def sale_return_cancel(request, return_id):
    sale_return = get_object_or_404(SaleReturn, pk=return_id)

    if sale_return.status != SaleReturn.Status.PENDENTE:
        messages.warning(request, "Esta devolução já foi processada.")
        return redirect('sale_detail', number=sale_return.sale.number)

    if request.method == 'POST':
        sale_return.status = SaleReturn.Status.CANCELADA
        sale_return.save()
        messages.success(request, f"Devolução #{sale_return.pk} cancelada.")

    return redirect('sale_detail', number=sale_return.sale.number)


# --- CONTAS A RECEBER (FINANCEIRO CENTRALIZADO) ---

@login_required
@user_passes_test(is_manager)
def billing_list(request):
    """Lista de todas as cobranças (Contas a Receber)."""
    billings = Billing.objects.all().select_related('client', 'sale', 'service_order')
    
    # Filtros simples
    status = request.GET.get('status')
    if status:
        billings = billings.filter(status=status)
        
    context = {
        'billings': billings,
        'title': 'Contas a Receber',
        'active_menu': 'finance'
    }
    return render(request, 'services/finance/billing_list.html', context)

@login_required
@user_passes_test(is_manager)
def billing_detail(request, pk):
    """Detalhamento de uma cobrança e suas parcelas."""
    billing = get_object_or_404(Billing, pk=pk)
    installments = billing.installments.all().select_related('payment_method')
    payment_methods = PaymentMethod.objects.filter(ativo=True)
    
    context = {
        'billing': billing,
        'installments': installments,
        'payment_methods': payment_methods,
        'title': f'Cobrança #{billing.number}'
    }
    return render(request, 'services/finance/billing_detail.html', context)

@login_required
@user_passes_test(is_manager)
def installment_pay(request, pk):
    """Dá baixa manual em uma parcela, suportando múltiplos meios de pagamento."""
    installment = get_object_or_404(Installment, pk=pk)
    
    if request.method == 'POST':
        # Recebe listas de métodos e valores do formulário
        method_ids = request.POST.getlist('payment_method[]')
        amounts = request.POST.getlist('payment_amount[]')
        
        if not method_ids:
            # Fallback para o comportamento antigo (único select)
            method_ids = [request.POST.get('payment_method')]
            amounts = [installment.amount - installment.get_total_paid()]
        
        total_this_time = Decimal('0.00')
        
        for m_id, amt, dt in zip(method_ids, amounts, request.POST.getlist('payment_date[]')):
            if not m_id or not amt: continue
            
            amt_decimal = Decimal(amt.replace(',', '.'))
            if amt_decimal <= 0: continue
            
            method = get_object_or_404(PaymentMethod, pk=m_id)
            
            pay_date = timezone.now()
            if dt:
                try:
                    pay_date = timezone.make_aware(datetime.fromisoformat(dt))
                except ValueError:
                    pass
            
            # Cria o registro de pagamento real
            SalePayment.objects.create(
                venda=installment.billing.sale,
                os=installment.billing.service_order,
                installment=installment,
                metodo_pagamento=method,
                valor_bruto=amt_decimal,
                valor_tarifa=Decimal('0.00'), # TODO: Calcular tarifa se necessário
                valor_liquido=amt_decimal,     # TODO: Calcular líquido
                data_pagamento=pay_date,
                data_previsao=timezone.now().date() # TODO: Usar prazo do método
            )
            total_this_time += amt_decimal
        
        # Atualiza status da Parcela
        total_paid = installment.get_total_paid()
        if total_paid >= installment.amount:
            installment.status = Installment.Status.PAGO
            installment.paid_at = timezone.now()
        elif total_paid > 0:
            installment.status = Installment.Status.PARCIAL
        
        installment.save()
        
        # Atualiza status da Cobrança (Billing)
        billing = installment.billing
        all_installments = billing.installments.all()
        
        # Verifica se todas as parcelas estão pagas
        if not all_installments.exclude(status=Installment.Status.PAGO).exists():
            billing.status = Billing.Status.PAGO
        elif all_installments.filter(status__in=[Installment.Status.PAGO, Installment.Status.PARCIAL]).exists():
            billing.status = Billing.Status.PARCIAL
        else:
            billing.status = Billing.Status.PENDENTE
            
        billing.save()
        
        messages.success(request, f"Pagamento de R$ {total_this_time} registrado para a parcela {installment.installment_number}!")
        
    return redirect('billing_detail', pk=installment.billing.id)

# --- CONTAS A PAGAR (DESPESAS) ---

@login_required
@user_passes_test(is_manager)
def expense_list(request):
    from datetime import timedelta, date as date_type
    q = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    category_id = request.GET.get('category', '').strip()

    # Atualiza automaticamente parcelas vencidas e não pagas para status ATRASADO
    today_auto = timezone.localdate()
    ExpenseInstallment.objects.filter(
        due_date__lt=today_auto,
        status__in=[ExpenseInstallment.Status.PENDENTE, ExpenseInstallment.Status.PARCIAL],
    ).update(status=ExpenseInstallment.Status.ATRASADO)

    installments = ExpenseInstallment.objects.select_related(
        'expense', 'expense__supplier', 'expense__category',
        'recurrence_rule', 'recurrence_rule__supplier'
    ).order_by('due_date')

    if q:
        installments = installments.filter(
            Q(expense__description__icontains=q) |
            Q(expense__supplier__name__icontains=q) |
            Q(expense__supplier__display_name__icontains=q) |
            Q(recurrence_rule__description__icontains=q) |
            Q(recurrence_rule__supplier__name__icontains=q) |
            Q(document_ref__icontains=q)
        )

    if status_filter:
        installments = installments.filter(status=status_filter)

    if date_from:
        try:
            installments = installments.filter(due_date__gte=date_type.fromisoformat(date_from))
        except ValueError:
            pass

    if date_to:
        try:
            installments = installments.filter(due_date__lte=date_type.fromisoformat(date_to))
        except ValueError:
            pass

    if category_id:
        installments = installments.filter(
            Q(expense__category_id=category_id) | Q(recurrence_rule__category_id=category_id)
        )

    today = timezone.localdate()

    # Totalizadores para o contexto (antes de paginar)
    overdue_qs = ExpenseInstallment.objects.filter(
        due_date__lt=today,
        status__in=[ExpenseInstallment.Status.PENDENTE, ExpenseInstallment.Status.ATRASADO, ExpenseInstallment.Status.PARCIAL],
    )
    overdue_count = overdue_qs.count()
    overdue_total = sum(inst.amount_remaining for inst in overdue_qs) if overdue_count else Decimal('0')

    paginator = Paginator(installments, 25)
    page = request.GET.get('page')
    installments_page = paginator.get_page(page)

    recurrence_rules = RecurrenceRule.objects.select_related('supplier', 'category').order_by('-created_at')
    bank_accounts = BankAccount.objects.filter(is_active=True).order_by('name')
    categories = FinancialCategory.objects.all().order_by('name')

    context = {
        'installments': installments_page,
        'recurrence_rules': recurrence_rules,
        'bank_accounts': bank_accounts,
        'categories': categories,
        'status_choices': ExpenseInstallment.Status.choices,
        'payment_method_choices': ExpenseInstallment.PaymentMethod.choices,
        'current_status': status_filter,
        'current_q': q,
        'current_date_from': date_from,
        'current_date_to': date_to,
        'current_category': category_id,
        'today': today,
        'warning_date': today + timedelta(days=7),
        'overdue_count': overdue_count,
        'overdue_total': overdue_total,
        'title': 'Contas a Pagar',
        'active_menu': 'finance'
    }
    return render(request, 'services/finance/expense_list.html', context)

@login_required
@user_passes_test(is_manager)
def expense_create(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        formset = ExpenseInstallmentFormSet(request.POST, prefix='installments')

        if form.is_valid():
            is_recurrent = form.cleaned_data.get('is_recurrent')

            if is_recurrent:
                with transaction.atomic():
                    rule = RecurrenceRule.objects.create(
                        supplier=form.cleaned_data['supplier'],
                        category=form.cleaned_data.get('category'),
                        description=form.cleaned_data['description'],
                        amount=form.cleaned_data['total_amount'],
                        frequency=form.cleaned_data['frequency'],
                        start_date=form.cleaned_data['issue_date'],
                        end_date=form.cleaned_data.get('end_date'),
                        created_by=request.user,
                    )
                    expense = form.save(commit=False)
                    expense.installments_count = 0
                    expense.recurrence_rule = rule
                    expense.save()
                messages.success(request, "Despesa recorrente criada com sucesso!")
                return redirect('expense_list')

            elif formset.is_valid():
                with transaction.atomic():
                    expense = form.save()
                    instances = formset.save(commit=False)
                    for instance in instances:
                        instance.expense = expense
                        instance.save()
                messages.success(request, "Despesa criada com sucesso!")
                return redirect('expense_list')
            else:
                logger.warning(f"ExpenseInstallmentFormSet errors: {formset.errors}")
        else:
            logger.warning(f"ExpenseForm errors: {form.errors}")
    else:
        form = ExpenseForm()
        formset = ExpenseInstallmentFormSet(queryset=ExpenseInstallment.objects.none(), prefix='installments')

    context = {
        'form': form,
        'formset': formset,
        'title': 'Nova Conta a Pagar',
        'is_edit': False,
        'active_menu': 'finance'
    }
    return render(request, 'services/finance/expense_form.html', context)

@login_required
@user_passes_test(is_manager)
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    rule = expense.recurrence_rule

    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        formset = ExpenseInstallmentFormSet(request.POST, instance=expense, prefix='installments')

        if form.is_valid():
            is_recurrent = form.cleaned_data.get('is_recurrent')

            if is_recurrent:
                with transaction.atomic():
                    if rule:
                        rule.supplier = form.cleaned_data['supplier']
                        rule.category = form.cleaned_data.get('category')
                        rule.description = form.cleaned_data['description']
                        rule.amount = form.cleaned_data['total_amount']
                        rule.frequency = form.cleaned_data['frequency']
                        rule.start_date = form.cleaned_data['issue_date']
                        rule.end_date = form.cleaned_data.get('end_date')
                        rule.save()
                    else:
                        rule = RecurrenceRule.objects.create(
                            supplier=form.cleaned_data['supplier'],
                            category=form.cleaned_data.get('category'),
                            description=form.cleaned_data['description'],
                            amount=form.cleaned_data['total_amount'],
                            frequency=form.cleaned_data['frequency'],
                            start_date=form.cleaned_data['issue_date'],
                            end_date=form.cleaned_data.get('end_date'),
                            created_by=request.user,
                        )
                    saved = form.save(commit=False)
                    saved.installments_count = 0
                    saved.recurrence_rule = rule
                    saved.save()
                messages.success(request, "Despesa recorrente atualizada com sucesso!")
                return redirect('expense_list')

            elif formset.is_valid():
                with transaction.atomic():
                    form.save()
                    formset.save()
                messages.success(request, "Despesa atualizada com sucesso!")
                return redirect('expense_list')
            else:
                logger.warning(f"ExpenseInstallmentFormSet errors: {formset.errors}")
    else:
        form = ExpenseForm(instance=expense)
        formset = ExpenseInstallmentFormSet(instance=expense, prefix='installments')

    context = {
        'form': form,
        'formset': formset,
        'title': f'Editar Despesa: {expense.description}',
        'is_edit': True,
        'active_menu': 'finance'
    }
    return render(request, 'services/finance/expense_form.html', context)

@login_required
@user_passes_test(is_manager)
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        if expense.recurrence_rule_id:
            messages.error(request, "Não é possível excluir uma despesa vinculada a uma recorrência. Desative a recorrência primeiro na aba Recorrências.")
            return redirect('expense_list')
        if expense.installments.exists():
            messages.error(request, "Não é possível excluir uma despesa que possui parcelas registradas.")
            return redirect('expense_list')
        expense.delete()
        messages.success(request, "Despesa excluída com sucesso!")
        return redirect('expense_list')
    return redirect('expense_list')

@login_required
@user_passes_test(is_manager)
def expense_installment_status_update(request, pk):
    installment = get_object_or_404(ExpenseInstallment, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        valid = {s[0] for s in ExpenseInstallment.Status.choices}
        if new_status in valid:
            installment.status = new_status
            if new_status == ExpenseInstallment.Status.PAGO and not installment.payment_date:
                installment.payment_date = timezone.localdate()
            installment.save()
            messages.success(request, f"Parcela marcada como {installment.get_status_display()}.")
    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('expense_list')


@login_required
@user_passes_test(is_manager)
def installment_payment_modal(request, pk, _payment_just_added=False, _discount_just_applied=False):
    installment = get_object_or_404(
        ExpenseInstallment.objects.select_related(
            'expense', 'expense__supplier',
            'recurrence_rule', 'recurrence_rule__supplier',
        ).prefetch_related('payments__bank_account', 'payments__created_by', 'payments__attachments'),
        pk=pk
    )
    bank_accounts = BankAccount.objects.filter(is_active=True).order_by('name')
    context = {
        'installment': installment,
        'bank_accounts': bank_accounts,
        'payment_method_choices': ExpenseInstallment.PaymentMethod.choices,
        'today': timezone.localdate(),
        'payment_just_added': _payment_just_added,
        'discount_just_applied': _discount_just_applied,
        'MEDIA_URL': '/media/',
    }
    return render(request, 'services/partials/installment_payment_modal.html', context)


@login_required
@user_passes_test(is_manager)
def installment_payment_add(request, pk):
    from decimal import InvalidOperation
    installment = get_object_or_404(ExpenseInstallment, pk=pk)
    if request.method == 'POST':
        amount_raw = request.POST.get('amount', '').strip().replace(',', '.')
        try:
            amount = Decimal(amount_raw)
            if amount <= 0:
                raise ValueError
        except (ValueError, InvalidOperation):
            messages.error(request, "Valor inválido.")
            if request.headers.get('HX-Request'):
                return installment_payment_modal(request, pk)
            return redirect('expense_list')

        payment_date_str = request.POST.get('payment_date', '')
        try:
            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            payment_date = timezone.localdate()

        payment_method = request.POST.get('payment_method') or None
        bank_account_id = request.POST.get('bank_account') or None
        notes = request.POST.get('notes', '')

        bank_account = None
        if bank_account_id:
            try:
                bank_account = BankAccount.objects.get(pk=bank_account_id, is_active=True)
            except BankAccount.DoesNotExist:
                pass

        payment_obj = InstallmentPayment.objects.create(
            installment=installment,
            amount=amount,
            payment_date=payment_date,
            payment_method=payment_method,
            bank_account=bank_account,
            notes=notes,
            created_by=request.user,
        )

        # Salva número de documento na parcela se informado
        document_ref = request.POST.get('document_ref', '').strip()
        if document_ref:
            installment.document_ref = document_ref
            installment.save(update_fields=['document_ref'])

        # Salva comprovante de pagamento se enviado
        attachment_file = request.FILES.get('attachment')
        if attachment_file:
            PaymentAttachment.objects.create(
                payment=payment_obj,
                file=attachment_file,
                description=request.POST.get('attachment_description', '').strip(),
            )

        installment.recalculate_status()

        if request.headers.get('HX-Request'):
            return installment_payment_modal(request, pk, _payment_just_added=True)

    if request.headers.get('HX-Request'):
        return installment_payment_modal(request, pk)

    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('expense_list')


@login_required
@user_passes_test(is_manager)
def installment_payment_delete(request, payment_pk):
    payment = get_object_or_404(InstallmentPayment, pk=payment_pk)
    installment = payment.installment
    if request.method == 'POST':
        payment.delete()
        installment.recalculate_status()
        messages.success(request, "Baixa removida com sucesso.")

    if request.headers.get('HX-Request'):
        return installment_payment_modal(request, installment.pk)

    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('expense_list')


@login_required
@user_passes_test(is_manager)
def expense_installments_export(request):
    """Exporta parcelas filtradas como planilha Excel."""
    qs = ExpenseInstallment.objects.select_related(
        'expense', 'expense__supplier', 'expense__category',
        'recurrence_rule', 'recurrence_rule__supplier',
    ).order_by('due_date')

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    category_id = request.GET.get('category', '').strip()

    if q:
        qs = qs.filter(
            Q(expense__supplier__display_name__icontains=q) |
            Q(expense__description__icontains=q) |
            Q(document_ref__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    if date_from:
        try:
            from datetime import date as date_type
            qs = qs.filter(due_date__gte=date_type.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import date as date_type
            qs = qs.filter(due_date__lte=date_type.fromisoformat(date_to))
        except ValueError:
            pass
    if category_id:
        qs = qs.filter(expense__category_id=category_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Vencimento", "Fornecedor", "Descrição", "Categoria", "Parcela", "Nº Doc.", "Valor (R$)", "Pago (R$)", "Saldo (R$)", "Status"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    col_widths = [14, 28, 32, 20, 10, 16, 14, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    status_labels = dict(ExpenseInstallment.Status.choices)
    for inst in qs:
        supplier = ''
        description = ''
        category = ''
        if inst.recurrence_rule:
            supplier = inst.recurrence_rule.supplier.display_name
            description = inst.recurrence_rule.description
            if inst.recurrence_rule.category:
                category = inst.recurrence_rule.category.name
        elif inst.expense:
            supplier = inst.expense.supplier.display_name
            description = inst.expense.description
            if inst.expense.category:
                category = inst.expense.category.name

        row = [
            inst.due_date.strftime('%d/%m/%Y'),
            supplier,
            description,
            category,
            inst.installment_label,
            inst.document_ref or '',
            float(inst.amount),
            float(inst.amount_paid),
            float(inst.amount_remaining),
            status_labels.get(inst.status, inst.status),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = border
        for col_idx in [7, 8, 9]:
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    # Totals row
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(float(i.amount) for i in qs)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=8, value=sum(float(i.amount_paid) for i in qs)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9, value=sum(float(i.amount_remaining) for i in qs)).number_format = '#,##0.00'
    for col_idx in [1, 7, 8, 9]:
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from django.utils.timezone import now
    filename = f"contas_a_pagar_{now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(is_manager)
def render_expense_installments_preview(request):
    try:
        data = request.POST if request.method == 'POST' else request.GET
        
        # Tenta pegar o valor de vários jeitos comuns
        total_amount_raw = data.get('total_amount', '0')
        # Limpa possíveis caracteres indesejados (espaços, R$)
        total_amount_raw = total_amount_raw.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
        try:
            total_amount = Decimal(total_amount_raw)
        except:
            total_amount = Decimal('0')
        
        installments_count_raw = data.get('installments_count', '1')
        try:
            installments_count = int(installments_count_raw)
        except:
            installments_count = 1
        
        frequency = data.get('frequency', '')
        issue_date_str = data.get('issue_date')
        
        logger.info(f"Parsed values: total={total_amount}, count={installments_count}, freq={frequency}, date={issue_date_str}")
        
        issue_date = timezone.now().date()
        if issue_date_str:
            try:
                issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        preview_data = []
        if installments_count > 0:
            # Evita divisão por zero
            div_count = Decimal(installments_count) if installments_count > 0 else Decimal('1')
            base_amount = (total_amount / div_count).quantize(Decimal('0.01'))
            remainder = total_amount - (base_amount * Decimal(installments_count))
            
            current_date = issue_date
            for i in range(1, installments_count + 1):
                amount = base_amount
                if i == installments_count:
                    amount += remainder
                    
                preview_data.append({
                    'installment_number': f"{i}/{installments_count}",
                    'amount': amount,
                    'due_date': current_date
                })
                
                if i < installments_count:
                    if frequency == Expense.Frequency.DIARIA:
                        current_date += relativedelta(days=1)
                    elif frequency == Expense.Frequency.SEMANAL:
                        current_date += relativedelta(days=7)
                    elif frequency == Expense.Frequency.QUINZENAL:
                        current_date += relativedelta(days=15)
                    elif frequency == Expense.Frequency.MENSAL:
                        current_date += relativedelta(months=1)
                    elif frequency == Expense.Frequency.ANUAL:
                        current_date += relativedelta(years=1)
                    else:
                        current_date += relativedelta(months=1)

        logger.info(f"Generated preview_data with {len(preview_data)} items")

        # Usar inlineformset_factory para total consistência com o que o save espera
        from django.forms import inlineformset_factory
        from .forms import ExpenseInstallmentForm
        
        DynamicFormSet = inlineformset_factory(
            Expense, ExpenseInstallment,
            form=ExpenseInstallmentForm,
            extra=installments_count,
            can_delete=False
        )
        
        formset = DynamicFormSet(
            queryset=ExpenseInstallment.objects.none(),
            initial=preview_data,
            prefix='installments'
        )
        logger.info(f"Formset initialized with {formset.total_form_count()} forms")
        
    except Exception as e:
        logger.exception(f"Erro no preview de parcelas: {str(e)}")
        # Fallback para o formset padrão definido no forms.py
        from .forms import ExpenseInstallmentFormSet
        formset = ExpenseInstallmentFormSet(queryset=ExpenseInstallment.objects.none(), prefix='installments')
        
    return render(request, 'services/partials/expense_installments_preview.html', {
        'formset': formset
    })


# --- FORECAST API ---

def _fmt_installment_label(num, rule=None, due_date=None):
    if num and '/' in num:
        parts = num.split('/')
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return f"{parts[0]} de {parts[1]}"
    if rule and due_date:
        end = rule.end_date or (rule.start_date + relativedelta(years=5))
        all_dates = rule.get_due_dates_in_range(rule.start_date, end)
        try:
            pos = all_dates.index(due_date) + 1
            total = len(all_dates)
            return f"{pos} de {total}" if rule.end_date else f"#{pos}"
        except ValueError:
            pass
    return num or '—'


@login_required
@user_passes_test(is_manager)
def expense_forecast_api(request):
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')

    if not start_str or not end_str:
        return JsonResponse({'error': 'Parâmetros start e end são obrigatórios'}, status=400)

    try:
        from datetime import date
        start = date.fromisoformat(start_str)
        end = date.fromisoformat(end_str)
    except ValueError:
        return JsonResponse({'error': 'Formato de data inválido. Use YYYY-MM-DD'}, status=400)

    # 1. Registros reais no período
    real_qs = ExpenseInstallment.objects.filter(
        due_date__range=(start, end)
    ).select_related('expense', 'expense__supplier', 'expense__category', 'recurrence_rule')

    real_data = []
    generated_keys = set()

    for inst in real_qs:
        supplier_name = ''
        description = ''
        if inst.recurrence_rule:
            supplier_name = inst.recurrence_rule.supplier.display_name
            description = inst.recurrence_rule.description
            generated_keys.add((inst.recurrence_rule_id, inst.recurrence_due_date))
        elif inst.expense:
            supplier_name = inst.expense.supplier.display_name
            description = inst.expense.description

        real_data.append({
            'id': inst.id,
            'type': 'real',
            'description': description,
            'supplier': supplier_name,
            'amount': str(inst.amount),
            'due_date': inst.due_date.isoformat(),
            'status': inst.status,
            'installment_number': inst.installment_number,
            'installment_label': _fmt_installment_label(
                inst.installment_number,
                rule=inst.recurrence_rule,
                due_date=inst.recurrence_due_date,
            ),
        })

    # 2. Projeções de regras de recorrência
    rules = RecurrenceRule.objects.filter(is_active=True).select_related('supplier', 'category')
    projected_data = []

    for rule in rules:
        rule_end = rule.end_date or (rule.start_date + relativedelta(years=5))
        all_rule_dates = rule.get_due_dates_in_range(rule.start_date, rule_end)
        due_dates = rule.get_due_dates_in_range(start, end)
        for due_date in due_dates:
            if (rule.id, due_date) not in generated_keys:
                try:
                    pos = all_rule_dates.index(due_date) + 1
                    total = len(all_rule_dates)
                    label = f"{pos} de {total}" if rule.end_date else f"#{pos}"
                except ValueError:
                    label = '—'
                projected_data.append({
                    'id': None,
                    'type': 'projected',
                    'description': rule.description,
                    'supplier': rule.supplier.display_name,
                    'amount': str(rule.amount),
                    'due_date': due_date.isoformat(),
                    'status': 'PREVISTO',
                    'installment_number': label,
                    'installment_label': label,
                    'rule_id': str(rule.id),
                })

    all_entries = sorted(real_data + projected_data, key=lambda x: x['due_date'])
    return JsonResponse({'results': all_entries, 'count': len(all_entries)})


# --- CONFIGURAÇÕES FINANCEIRAS ---

@login_required
@user_passes_test(is_manager)
def finance_settings_view(request):
    obj, _ = FinanceSettings.objects.get_or_create(pk=1)
    form = FinanceSettingsForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Configurações financeiras salvas.')
        return redirect('finance_settings')
    return render(request, 'services/finance/finance_settings.html', {
        'form': form,
        'title': 'Configurações Financeiras',
        'active_menu': 'finance',
    })


@login_required
@user_passes_test(is_manager)
def sale_settings_view(request):
    obj = SaleSettings.get()
    form = SaleSettingsForm(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Configurações de vendas salvas com sucesso.')
        return redirect('sale_settings')
    return render(request, 'services/sale_settings.html', {
        'form': form,
        'settings': obj,
        'active_menu': 'sales',
    })


# --- RECORRÊNCIAS ---

@login_required
@user_passes_test(is_manager)
def recurrence_rule_toggle(request, pk):
    rule = get_object_or_404(RecurrenceRule, pk=pk)
    if request.method == 'POST':
        rule.is_active = not rule.is_active
        rule.save()
        status = 'ativada' if rule.is_active else 'desativada'
        messages.success(request, f'Recorrência {status} com sucesso.')
    return redirect('expense_list')


@login_required
@user_passes_test(is_manager)
def recurrence_materialize(request, rule_pk):
    """Materializa antecipadamente uma parcela prevista de uma recorrência."""
    if request.method != 'POST':
        return redirect('expense_list')

    rule = get_object_or_404(RecurrenceRule, pk=rule_pk, is_active=True)
    due_date_str = request.POST.get('due_date', '')

    try:
        from datetime import date as date_type
        due_date = date_type.fromisoformat(due_date_str)
    except (ValueError, TypeError):
        messages.error(request, 'Data de vencimento inválida.')
        return redirect('expense_list')

    rule_end = rule.end_date or (rule.start_date + relativedelta(years=5))
    valid_dates = rule.get_due_dates_in_range(rule.start_date, rule_end)
    if due_date not in valid_dates:
        messages.error(request, 'A data informada não é uma ocorrência válida desta recorrência.')
        return redirect('expense_list')

    existing = ExpenseInstallment.objects.filter(
        recurrence_rule=rule,
        recurrence_due_date=due_date,
    ).first()

    if existing:
        messages.info(request, 'Esta parcela já foi gerada anteriormente.')
        if request.headers.get('HX-Request'):
            return installment_payment_modal(request, existing.pk)
        return redirect('expense_list')

    template_expense = rule.expense_templates.first()
    if not template_expense:
        messages.error(request, 'Nenhuma despesa modelo associada a esta recorrência.')
        return redirect('expense_list')

    installment = ExpenseInstallment.objects.create(
        expense=template_expense,
        recurrence_rule=rule,
        recurrence_due_date=due_date,
        installment_number=due_date.strftime('%m/%Y'),
        due_date=due_date,
        amount=rule.amount,
        status=ExpenseInstallment.Status.PENDENTE,
    )

    if request.headers.get('HX-Request'):
        return installment_payment_modal(request, installment.pk)

    messages.success(request, f'Parcela de {rule.description} ({due_date.strftime("%d/%m/%Y")}) gerada com sucesso.')
    return redirect('expense_list')


@login_required
@user_passes_test(is_manager)
def installment_apply_discount(request, pk):
    """Quita uma parcela por desconto/isenção sem registrar pagamento real."""
    if request.method != 'POST':
        return redirect('expense_list')

    installment = get_object_or_404(ExpenseInstallment, pk=pk)

    if installment.amount_remaining <= Decimal('0'):
        if request.headers.get('HX-Request'):
            return installment_payment_modal(request, pk)
        return redirect('expense_list')

    discount_reason = request.POST.get('discount_reason', '').strip()
    if not discount_reason:
        if request.headers.get('HX-Request'):
            return installment_payment_modal(request, pk)
        messages.error(request, 'Informe o motivo do desconto/isenção.')
        return redirect('expense_list')

    from decimal import InvalidOperation
    amount_str = request.POST.get('amount', '').replace(',', '.').strip()
    try:
        amount = Decimal(amount_str) if amount_str else installment.amount_remaining
        if amount <= Decimal('0') or amount > installment.amount_remaining:
            raise ValueError
    except (ValueError, InvalidOperation):
        amount = installment.amount_remaining

    InstallmentPayment.objects.create(
        installment=installment,
        amount=amount,
        payment_date=timezone.localdate(),
        is_discount=True,
        discount_reason=discount_reason,
        created_by=request.user,
    )
    installment.recalculate_status()

    if request.headers.get('HX-Request'):
        return installment_payment_modal(request, pk, _discount_just_applied=True)
    return redirect('expense_list')


@login_required
@user_passes_test(is_manager)
def expense_cashflow(request):
    """Página de Fluxo de Caixa: saldo atual das contas + projeção de entradas e saídas."""
    from dateutil.relativedelta import relativedelta
    from django.db.models import Sum
    from decimal import Decimal
    import calendar

    today = timezone.localdate()
    months_ahead = 3

    # Período de análise: mês atual + próximos 3 meses
    period_start = today.replace(day=1)
    period_end = (period_start + relativedelta(months=months_ahead + 1)) - relativedelta(days=1)

    # Contas bancárias com saldo atual
    bank_accounts = BankAccount.objects.filter(is_active=True).order_by('name')
    total_bank_balance = Decimal('0')
    for acc in bank_accounts:
        total_bank_balance += acc.current_balance

    # Saídas previstas por mês (parcelas a pagar)
    outflows_qs = ExpenseInstallment.objects.filter(
        due_date__range=(period_start, period_end),
        status__in=[ExpenseInstallment.Status.PENDENTE, ExpenseInstallment.Status.PARCIAL, ExpenseInstallment.Status.ATRASADO],
    ).select_related('expense', 'expense__supplier', 'recurrence_rule', 'recurrence_rule__supplier')

    # Entradas previstas por mês (contas a receber - Installments pendentes)
    try:
        from .models import Installment as BillingInstallment
        inflows_qs = BillingInstallment.objects.filter(
            due_date__range=(period_start, period_end),
            status__in=['PENDENTE', 'PARCIAL', 'ATRASADO'],
        )
        has_billing = True
    except Exception:
        inflows_qs = []
        has_billing = False

    # Monta estrutura mensal
    months = []
    for i in range(months_ahead + 1):
        month_start = period_start + relativedelta(months=i)
        month_end = (month_start + relativedelta(months=1)) - relativedelta(days=1)

        month_outflows = [
            inst for inst in outflows_qs
            if month_start <= inst.due_date <= month_end
        ]
        month_inflows = []
        if has_billing:
            month_inflows = [
                inst for inst in inflows_qs
                if month_start <= inst.due_date <= month_end
            ]

        total_out = sum(inst.amount_remaining for inst in month_outflows)
        total_in = sum(
            (inst.amount - (inst.amount_paid if hasattr(inst, 'amount_paid') else Decimal('0')))
            for inst in month_inflows
        ) if month_inflows else Decimal('0')

        months.append({
            'label': month_start.strftime('%B/%Y'),
            'month_start': month_start,
            'month_end': month_end,
            'outflows': month_outflows[:20],
            'inflows': month_inflows[:20],
            'total_out': total_out,
            'total_in': total_in,
            'net': total_in - total_out,
        })

    # Parcelas atrasadas
    overdue = ExpenseInstallment.objects.filter(
        due_date__lt=today,
        status__in=[ExpenseInstallment.Status.PENDENTE, ExpenseInstallment.Status.PARCIAL, ExpenseInstallment.Status.ATRASADO],
    ).select_related('expense', 'expense__supplier', 'recurrence_rule', 'recurrence_rule__supplier').order_by('due_date')[:15]

    overdue_total = sum(inst.amount_remaining for inst in overdue)

    context = {
        'bank_accounts': bank_accounts,
        'total_bank_balance': total_bank_balance,
        'months': months,
        'today': today,
        'overdue': overdue,
        'overdue_total': overdue_total,
        'has_billing': has_billing,
        'active_menu': 'finance',
        'title': 'Fluxo de Caixa',
    }
    return render(request, 'services/finance/cashflow.html', context)


@login_required
@user_passes_test(is_manager)
def bank_account_list(request):
    accounts = BankAccount.objects.all().order_by('name')
    context = {
        'accounts': accounts,
        'title': 'Contas Bancárias',
        'active_menu': 'finance',
    }
    return render(request, 'services/finance/bank_accounts.html', context)


@login_required
@user_passes_test(is_manager)
def bank_account_create(request):
    from django import forms as dj_forms

    class BankAccountForm(dj_forms.ModelForm):
        class Meta:
            model = BankAccount
            fields = ['name', 'bank', 'account_type', 'initial_balance', 'is_active']

    if request.method == 'POST':
        form = BankAccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta bancária criada com sucesso.')
            return redirect('bank_account_list')
    else:
        form = BankAccountForm()

    return render(request, 'services/finance/bank_account_form.html', {
        'form': form,
        'title': 'Nova Conta Bancária',
        'is_edit': False,
        'active_menu': 'finance',
    })


@login_required
@user_passes_test(is_manager)
def bank_account_edit(request, pk):
    from django import forms as dj_forms
    account = get_object_or_404(BankAccount, pk=pk)

    class BankAccountForm(dj_forms.ModelForm):
        class Meta:
            model = BankAccount
            fields = ['name', 'bank', 'account_type', 'initial_balance', 'is_active']

    if request.method == 'POST':
        form = BankAccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta bancária atualizada.')
            return redirect('bank_account_list')
    else:
        form = BankAccountForm(instance=account)

    return render(request, 'services/finance/bank_account_form.html', {
        'form': form,
        'title': f'Editar: {account.name}',
        'is_edit': True,
        'active_menu': 'finance',
        'account': account,
    })


@login_required
@user_passes_test(is_manager)
def bank_account_toggle(request, pk):
    account = get_object_or_404(BankAccount, pk=pk)
    if request.method == 'POST':
        account.is_active = not account.is_active
        account.save(update_fields=['is_active'])
        status = 'ativada' if account.is_active else 'desativada'
        messages.success(request, f'Conta "{account.name}" {status}.')
    return redirect('bank_account_list')


@login_required
@user_passes_test(is_manager)
def payment_attachment_delete(request, pk):
    attachment = get_object_or_404(PaymentAttachment, pk=pk)
    installment_pk = attachment.payment.installment_id
    if request.method == 'POST':
        attachment.file.delete(save=False)
        attachment.delete()
        messages.success(request, 'Comprovante removido.')
    if request.headers.get('HX-Request'):
        return installment_payment_modal(request, installment_pk)
    return redirect('expense_list')

