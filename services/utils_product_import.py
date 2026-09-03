import csv
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import Product


BLING_REQUIRED_HEADERS = {'id', 'codigo', 'descricao', 'preco', 'estoque'}


def normalize_header(value):
    text = unicodedata.normalize('NFKD', str(value or ''))
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return re.sub(r'[^a-z0-9]+', '_', text.lower()).strip('_')


def clean_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace('\t', '').strip()


def decimal_from_brazilian(value, default='0'):
    text = clean_text(value)
    if not text:
        return Decimal(default)
    if ',' in text:
        text = text.replace('.', '').replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(f'Número inválido: {value}')


def digits_only(value):
    return re.sub(r'\D', '', clean_text(value))


def _json_value(value):
    if value in (None, ''):
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def read_spreadsheet(file_content, filename):
    extension = filename.lower().rsplit('.', 1)[-1]
    if extension == 'csv':
        try:
            decoded = file_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = file_content.decode('cp1252')
        reader = csv.DictReader(io.StringIO(decoded))
        return list(reader)

    if extension == 'xlsx':
        workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, ())
        return [dict(zip(headers, values)) for values in rows]

    if extension == 'xls':
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError('Suporte a XLS indisponível. Instale a dependência xlrd.') from exc
        workbook = xlrd.open_workbook(file_contents=file_content)
        sheet = workbook.sheet_by_index(0)
        if not sheet.nrows:
            return []
        headers = sheet.row_values(0)
        return [dict(zip(headers, sheet.row_values(index))) for index in range(1, sheet.nrows)]

    raise ValueError('Formato não suportado. Envie CSV, XLSX ou XLS.')


UNIT_MAP = {
    'UN': Product.UnitType.UNIDADE,
    'MT': Product.UnitType.METRO,
    'M': Product.UnitType.METRO,
    'M2': Product.UnitType.M2,
    'L': Product.UnitType.LITRO,
    'KG': Product.UnitType.QUILO,
    'PC': Product.UnitType.PECA,
    'CX': Product.UnitType.CAIXA,
    'CE': Product.UnitType.CENTO,
    'CT': Product.UnitType.CARTELA,
    'PR': Product.UnitType.PAR,
    'PCT': Product.UnitType.PACOTE,
    'PT': Product.UnitType.POTE,
    'RL': Product.UnitType.ROLO,
    'MO': Product.UnitType.MAO_OBRA,
    'BR': Product.UnitType.BARRA,
    'JO': Product.UnitType.JOGO,
    'HR': Product.UnitType.HORA,
    'TB': Product.UnitType.TUBO,
    'SC': Product.UnitType.SACO,
}

TYPE_MAP = {
    'servicos': Product.Type.SERVICO,
    'produto_acabado': Product.Type.PRODUTO_ACABADO,
    'materia_prima': Product.Type.MATERIA_PRIMA,
}


def _dimension_in_cm(value, unit):
    amount = decimal_from_brazilian(value)
    normalized_unit = normalize_header(unit)
    if normalized_unit == 'milimetro':
        return amount / Decimal('10')
    if normalized_unit in {'metro', 'metros'}:
        return amount * Decimal('100')
    return amount


def parse_bling_rows(file_content, filename):
    raw_rows = read_spreadsheet(file_content, filename)
    if not raw_rows:
        raise ValueError('A planilha está vazia.')

    available_headers = {normalize_header(header) for header in raw_rows[0]}
    missing = BLING_REQUIRED_HEADERS - available_headers
    if missing:
        raise ValueError(f"Planilha do Bling inválida. Colunas ausentes: {', '.join(sorted(missing))}.")

    parsed_rows = []
    for row_number, raw in enumerate(raw_rows, start=2):
        values = {normalize_header(header): value for header, value in raw.items()}
        if not any(clean_text(value) for value in raw.values()):
            continue

        row_errors = []
        try:
            bling_id_text = clean_text(values.get('id'))
            bling_id = int(Decimal(bling_id_text))
        except (InvalidOperation, ValueError):
            bling_id = None
            row_errors.append('ID do Bling inválido.')

        name = clean_text(values.get('descricao'))
        bling_code = clean_text(values.get('codigo'))
        if not name:
            row_errors.append('Descrição obrigatória.')
        if not bling_code:
            row_errors.append('Código do Bling obrigatório.')

        def number(key, default='0'):
            try:
                return str(decimal_from_brazilian(values.get(key), default))
            except ValueError as exc:
                row_errors.append(f'{key}: {exc}')
                return default

        raw_external = {
            clean_text(header): _json_value(value)
            for header, value in raw.items()
            if clean_text(header) and clean_text(value)
        }

        situation = normalize_header(values.get('situacao'))
        item_type = normalize_header(values.get('tipo_do_item'))
        dimension_unit = values.get('unidade_de_medida')
        try:
            width = str(_dimension_in_cm(values.get('largura_do_produto'), dimension_unit))
            height = str(_dimension_in_cm(values.get('altura_do_produto'), dimension_unit))
            depth = str(_dimension_in_cm(values.get('profundidade_do_produto'), dimension_unit))
        except ValueError as exc:
            width = height = depth = '0'
            row_errors.append(f'Dimensões: {exc}')

        origin_text = clean_text(values.get('origem'))
        try:
            origin = int(Decimal(origin_text or '0'))
            if origin not in range(9):
                raise ValueError
        except (InvalidOperation, ValueError):
            origin = 0
            row_errors.append('Origem da mercadoria inválida.')

        unit_code = clean_text(values.get('unidade')).upper()
        parsed_rows.append({
            'row_number': row_number,
            'bling_id': bling_id,
            'bling_code': bling_code,
            'name': name,
            'supplier_code': clean_text(values.get('cod_no_fornecedor')),
            'supplier_name': clean_text(values.get('fornecedor')),
            'category_name': clean_text(values.get('categoria_do_produto')),
            'barcode': clean_text(values.get('gtin_ean')) or None,
            'unit_type': UNIT_MAP.get(unit_code, Product.UnitType.UNIDADE),
            'unit_original': unit_code,
            'type': TYPE_MAP.get(item_type, Product.Type.PRODUTO),
            'format': Product.Format.SIMPLES,
            'default_unit_price': number('preco'),
            'preco_custo': number('preco_de_custo'),
            'stock': number('estoque'),
            'min_stock': number('estoque_minimo'),
            'max_stock': number('estoque_maximo'),
            'weight': number('peso_bruto_kg'),
            'net_weight': number('peso_liquido_kg'),
            'width': width,
            'height': height,
            'depth': depth,
            'ncm': digits_only(values.get('ncm')) or None,
            'cest': digits_only(values.get('cest')) or None,
            'origem_mercadoria': origin,
            'is_active': situation != 'inativo',
            'bling_product_kind': clean_text(values.get('produto_variacao')),
            'external_data': {'bling': raw_external},
            'errors': row_errors,
        })

    return parsed_rows
